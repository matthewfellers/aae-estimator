import os, json, base64, re, hashlib
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, g, after_this_request
from anthropic import Anthropic
from supabase import create_client
import io
from datetime import datetime
from scanner_pipeline import scan_drawing as _pipeline_scan_drawing
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Security imports ──────────────────────────────────────────────────────────
from security.auth import require_auth, require_admin, require_role, get_user_sb

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", os.urandom(32))

# ── Rate limiting ─────────────────────────────────────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=[],
        storage_uri=os.environ.get("REDIS_URL", "memory://"),
    )
    _rate_limiting_available = True
except Exception:
    _rate_limiting_available = False
    limiter = None

# ── Security headers (applied to every response) ─────────────────────────────
@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Strict-Transport-Security"] = "max-age=63072000; includeSubDomains"
    # CSP: allow Supabase JS CDN, fonts, inline styles (needed for our SPA)
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: blob:; "
        "connect-src 'self' https://*.supabase.co https://api.anthropic.com; "
        "frame-ancestors 'none';"
    )
    response.headers["Content-Security-Policy"] = csp
    return response

# ── Clients ───────────────────────────────────────────────────────────────────
claude_client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
_sb_url = os.environ.get("SUPABASE_URL", "")
_sb_key = os.environ.get("SUPABASE_ANON_KEY", "")
# IMPORTANT: This global client uses the ANON key and runs WITHOUT a user token.
# It is intentionally used ONLY for the internal calculation helpers (get_labor_rates,
# get_vendor_map) that read reference data, NOT for any user-facing route handlers.
# All route handlers must use get_user_sb() so RLS evaluates as the actual user.
supabase = create_client(_sb_url, _sb_key) if _sb_url and _sb_key else None

# ── Audit logging ─────────────────────────────────────────────────────────────
def audit_log(action: str, entity: str = None, entity_id: str = None, payload: dict = None):
    """
    Write an audit record using the USER-SCOPED Supabase client.
    This is required for the RLS INSERT policy (auth.uid() = actor_user_id) to pass.
    Non-blocking best-effort — never raises.
    """
    if not _sb_url or not _sb_key:
        return
    try:
        user = getattr(g, "user", {})
        record = {
            "actor_user_id": user.get("id"),
            "actor_email":   user.get("email"),
            "org_id":        user.get("org_id"),   # required for org-scoped audit_log RLS
            "action":        action,
            "entity":        entity,
            "entity_id":     str(entity_id) if entity_id is not None else None,
            "payload":       payload,
        }
        # Must use user-scoped client so RLS passes auth.uid() = actor_user_id
        sb = get_user_sb()
        sb.table("audit_log").insert(record).execute()
    except Exception as e:
        print(f"[audit_log] write failed: {e}", flush=True)

# ── Labor rates (minutes per unit) ────────────────────────────────────────
# Labor rates in MINUTES per unit — calibrated against AAE actual job history
# Baseline: CR Permian PLC panel = 28 hrs actual
LABOR_RATES = {
    # Enclosure & mechanical
    "enclosure_prep": 20, "subpanel_mount": 15, "panel_layout": 25,
    "din_rail": 6, "wire_duct": 4, "enc_accessory": 8, "door_component": 12,
    # Power distribution
    "main_breaker_small": 15, "main_breaker_large": 30,
    "branch_breaker_1p": 6, "branch_breaker_23p": 10,
    "fused_disconnect": 15, "cpt": 20, "pdb": 12,
    # Motor control
    "relay_icecube": 8, "relay_din": 6, "contactor_small": 15,
    "contactor_large": 25, "overload": 10, "timer": 8, "ssr": 10,
    # VFDs / soft starters
    "vfd_small": 40, "vfd_med": 65, "vfd_large": 100,
    "soft_starter_small": 50, "soft_starter_large": 80,
    # Control devices
    "pilot_light": 6, "selector": 8, "pushbutton": 6, "estop": 12,
    # PLC / networking
    "plc_rack": 35, "plc_di_do": 2.5, "plc_ai_ao": 3.5, "hmi": 30,
    "safety_relay": 20, "eth_switch": 15, "eth_cable": 5,
    # Terminal blocks
    "tb_standard": 2.5, "tb_ground": 2.5, "tb_fused": 4, "tb_disconnect": 4,
    "tb_accessories": 3, "terminal_markers": 3,
    # Wiring (per wire, includes routing + landing both ends)
    "wire_land_control": 1.2, "ferrule": 0.25, "wire_route": 0.35,
    "heat_shrink_label": 0.25, "heat_shrink_batch": 2,
    # UL / QC (realistic shop times)
    "ul_labels": 15, "continuity_check": 0.4, "hipot": 12,
    "as_built": 25, "qc_signoff": 12,
}

COMPLEXITY_MULT = {
    "SIMPLE": 0.85, "STANDARD": 1.0, "MULTI": 1.1,
    "PLC-MOD": 1.15, "PLC-HIGH": 1.25, "PLC-DRIV": 1.35,
    "SAFETY": 1.4, "CUSTOM": 1.5,
}
TECH_MULT = {
    "MASTER": 0.85, "JOURNEYMAN": 1.0, "APPRENTICE": 1.3, "MIXED": 1.1,
}
SHOP_RATES = {
    "labor_rate": 95.0, "mat_markup": 0.30,
    "overhead": 0.12, "profit": 0.18, "expedite": 0.15,
    # Wire pricing per foot (AAE actual cost)
    "wire_10awg_under": 0.40, "wire_8awg": 0.65,
    "wire_6awg": 1.10, "wire_4awg": 1.75, "wire_2awg": 3.00,
    "wire_1awg": 3.50, "wire_1_0": 5.25, "wire_2_0": 5.50,
    "wire_3_0": 6.36, "wire_4_0": 7.50, "wire_250mcm": 13.00,
    # Heat shrink: H075X044H1T — $275/roll, 1000 labels, ~850 usable (2 rows wasted)
    "heat_shrink_roll_cost": 275.0,
    "heat_shrink_per_roll": 850,   # usable labels per roll
    # Legacy fallbacks
    "wire_16_per_ft": 0.40, "wire_14_per_ft": 0.40,
    "wire_12_per_ft": 0.40, "wire_10_per_ft": 0.40,
    "heat_shrink_each": round(275.0 / 850, 4),
}

# Wire gauge cost lookup (value from select = cost per foot)
WIRE_GAUGE_COSTS = {
    "10_under": 0.40, "8": 0.65, "6": 1.10, "4": 1.75,
    "2": 3.00, "1": 3.50, "1_0": 5.25, "2_0": 5.50,
    "3_0": 6.36, "4_0": 7.50, "250mcm": 13.00,
}

# ── Calculation engine ─────────────────────────────────────────────────────
def get_labor_rates():
    """Load labor rates from DB if available, fall back to defaults."""
    if not supabase:
        return LABOR_RATES.copy()
    try:
        rows = supabase.table("aae_labor_rates").select("rate_key,rate_value").execute()
        if rows.data:
            merged = LABOR_RATES.copy()
            for row in rows.data:
                merged[row["rate_key"]] = float(row["rate_value"])
            return merged
    except Exception:
        pass
    return LABOR_RATES.copy()

def get_vendor_map():
    """Load manufacturer->vendor mapping from DB."""
    vendor_map = {
        # defaults if DB unavailable
        "allen bradley": "Rexel", "rockwell automation": "Rexel",
        "hammond power": "Rexel", "panduit": "Rexel", "mersen": "Rexel",
        "hoffman": "Rexel", "n-tron": "Rexel", "corning": "Rexel",
        "phoenix contact": "AWC", "rittal": "AWC", "hammond enclosures": "AWC",
        "siemens": "AWC", "solar shield": "AWC", "bussmann": "AWC",
        "marathon special products": "AWC", "tripp lite": "AWC",
        "turck": "A-Tech", "red lion": "A-Tech",
        "square d": "Graybar", "schneider electric": "Graybar",
        "cisco": "TD Synnex",
        "saginaw control engineering": "Saginaw", "saginaw": "Saginaw", "sce": "Saginaw",
        "automation direct": "Automation Direct",
        "factorymation": "Factorymation",
    }
    if not supabase:
        return vendor_map
    try:
        rows = supabase.table("aae_vendors").select("manufacturer,vendor_name").eq("active", True).execute()
        if rows.data:
            for row in rows.data:
                vendor_map[row["manufacturer"].lower()] = row["vendor_name"]
    except Exception:
        pass
    return vendor_map


# ── Deterministic Vendor Routing Engine ──────────────────────────────────────
# 4-tier precedence: Part Override → Prefix Rule → Manufacturer Default → UNASSIGNED

# Hardcoded fallback defaults (from AAE_Vendor_Routing_Rules.json)
_DEFAULT_PART_OVERRIDES = {
    "ATQR1/2":       {"manufacturer": "Eaton Bussmann",    "vendor": "Rexel",     "note": "Purchasing exception (buy from Rexel)"},
    "ATQR1/4":       {"manufacturer": "Eaton Bussmann",    "vendor": "Rexel",     "note": "Purchasing exception (buy from Rexel)"},
    "CEP-FS30CC2":   {"manufacturer": "Mersen",            "vendor": "Rexel",     "note": "Mersen via Rexel"},
    "XETA9X-11INDFD":{"manufacturer": "XetaWave",         "vendor": "Rexel",     "note": "XetaWave via Rexel"},
    "A-4AXFN24":     {"manufacturer": "nVent HOFFMAN",    "vendor": "Rexel",     "note": ""},
    "P-R2-K2RF0":    {"manufacturer": "GracePort",        "vendor": "Rexel",     "note": ""},
    "STP480D07M":    {"manufacturer": "Mersen",            "vendor": "Rexel",     "note": ""},
    "STZ480D20B1":   {"manufacturer": "Mersen",            "vendor": "Rexel",     "note": ""},
    "WF100LP":       {"manufacturer": "nVent HOFFMAN",    "vendor": "Rexel",     "note": ""},
    "MB3170":        {"manufacturer": "MOXA",              "vendor": "AWC",       "note": ""},
    "MB3170-T":      {"manufacturer": "MOXA",              "vendor": "AWC",       "note": ""},
    "TRS15":         {"manufacturer": "Eaton Bussmann",    "vendor": "AWC",       "note": ""},
    "TRS40":         {"manufacturer": "Eaton Bussmann",    "vendor": "AWC",       "note": ""},
    "G07S0000":      {"manufacturer": "Red Lion Controls", "vendor": "Atech",     "note": ""},
    "G10S0000":      {"manufacturer": "Red Lion Controls", "vendor": "Atech",     "note": ""},
    "G15C0000":      {"manufacturer": "Red Lion Controls", "vendor": "Atech",     "note": ""},
    "DRP-03":        {"manufacturer": "Mean Well",         "vendor": "RS America","note": ""},
}

_DEFAULT_PREFIX_RULES_RAW = [
    # ── Schneider Electric / Square D ────────────────────────────────────────
    {"prefix": "QO",     "pattern": "^QO",                                       "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Schneider breakers",  "priority": 20},
    {"prefix": "9070TF", "pattern": "^9070TF",                                   "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Schneider xfmrs",     "priority": 60},
    {"prefix": "LV",     "pattern": "^LV\\d+",                                   "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Schneider LV",        "priority": 20},
    {"prefix": "BGA/BJA/HJA/JJA/RJA", "pattern": "^(BGA|BJA|HJA|JJA|RJA)\\d+", "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Schneider breakers",  "priority": 30},
    {"prefix": "NQ",     "pattern": "^NQ\\d+",                                   "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Schneider panelboards","priority": 20},
    {"prefix": "HOM",    "pattern": "^HOM",                                      "manufacturer": "Schneider Electric",                "vendor": "Graybar", "note": "Homeline breakers",   "priority": 20},

    # ── Rockwell Automation / Allen-Bradley ──────────────────────────────────
    # PLC / I/O modules
    {"prefix": "1769-",  "pattern": "^1769-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "CompactLogix I/O",    "priority": 50},
    {"prefix": "1756-",  "pattern": "^1756-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "ControlLogix",        "priority": 50},
    {"prefix": "1734-",  "pattern": "^1734-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "POINT I/O",           "priority": 50},
    {"prefix": "5034-",  "pattern": "^5034-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Flex 5000",           "priority": 50},
    {"prefix": "1794-",  "pattern": "^1794-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "FLEX I/O",            "priority": 50},
    {"prefix": "5094-",  "pattern": "^5094-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "FLEX 5000",           "priority": 50},
    {"prefix": "5069-",  "pattern": "^5069-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Compact 5000",        "priority": 50},
    {"prefix": "2080-",  "pattern": "^2080-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Micro800",            "priority": 50},
    # Networking
    {"prefix": "1783-",  "pattern": "^1783-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Stratix switches",    "priority": 50},
    # VFDs / Drives
    {"prefix": "20F",    "pattern": "^20F",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 753",       "priority": 30},
    {"prefix": "20G",    "pattern": "^20G",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 755",       "priority": 30},
    {"prefix": "20P",    "pattern": "^20P",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 700",       "priority": 30},
    {"prefix": "20A",    "pattern": "^20A",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 70",        "priority": 30},
    {"prefix": "20B",    "pattern": "^20B",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 700L",      "priority": 30},
    {"prefix": "20D",    "pattern": "^20D",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 700S",      "priority": 30},
    {"prefix": "22B",    "pattern": "^22B",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 40",        "priority": 30},
    {"prefix": "22C",    "pattern": "^22C",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 400",       "priority": 30},
    {"prefix": "22D",    "pattern": "^22D",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 40P",       "priority": 30},
    {"prefix": "25A",    "pattern": "^25A",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 523",       "priority": 30},
    {"prefix": "25B",    "pattern": "^25B",                                      "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex 525",       "priority": 30},
    # Drive accessories
    {"prefix": "20-HIM", "pattern": "^20-HIM",                                   "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex HIM",       "priority": 60},
    {"prefix": "20-750", "pattern": "^20-750",                                   "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex comm/opt",  "priority": 60},
    {"prefix": "20-COMM","pattern": "^20-COMM",                                  "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PowerFlex comms",     "priority": 60},
    # Breakers / Protection
    {"prefix": "140G-",  "pattern": "^140G-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Molded case breakers","priority": 50},
    {"prefix": "140M-",  "pattern": "^140M-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Motor protectors",    "priority": 50},
    {"prefix": "140MG-", "pattern": "^140MG-",                                   "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Motor protectors",    "priority": 60},
    {"prefix": "150-",   "pattern": "^150-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "SMC soft starters",   "priority": 50},
    # Terminal blocks / Wiring
    {"prefix": "1492-",  "pattern": "^1492-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Terminal blocks",     "priority": 50},
    # Transformers
    {"prefix": "1497-",  "pattern": "^1497-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Control transformers","priority": 50},
    # Contactors / Relays
    {"prefix": "100-",   "pattern": "^100-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Contactors",          "priority": 50},
    {"prefix": "100S-",  "pattern": "^100S-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Safety contactors",   "priority": 60},
    {"prefix": "700-",   "pattern": "^700-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Relays",              "priority": 50},
    {"prefix": "700S-",  "pattern": "^700S-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Safety relays",       "priority": 60},
    # Pushbuttons / Pilot devices
    {"prefix": "800F",   "pattern": "^800F",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "22mm pilot devices",  "priority": 40},
    {"prefix": "800T",   "pattern": "^800T",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "30mm pilot devices",  "priority": 40},
    {"prefix": "800H",   "pattern": "^800H",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "NEMA pilot devices",  "priority": 40},
    # HMI / Displays
    {"prefix": "2711R-", "pattern": "^2711R-",                                   "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PanelView 800",       "priority": 50},
    {"prefix": "2711P-", "pattern": "^2711P-",                                   "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "PanelView Plus",      "priority": 50},
    {"prefix": "6189-",  "pattern": "^6189-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Industrial computers","priority": 50},
    # Safety
    {"prefix": "440R-",  "pattern": "^440R-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Safety relays",       "priority": 50},
    {"prefix": "440G-",  "pattern": "^440G-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Safety interlocks",   "priority": 50},
    {"prefix": "440N-",  "pattern": "^440N-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Safety switches",     "priority": 50},
    # Overloads / Starters
    {"prefix": "193-",   "pattern": "^193-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Overload relays",     "priority": 50},
    {"prefix": "509-",   "pattern": "^509-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "NEMA starters",       "priority": 50},
    # DIN Rail / Accessories
    {"prefix": "199-",   "pattern": "^199-",                                     "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "DIN rail/accessories","priority": 50},
    # Power supplies
    {"prefix": "1606-",  "pattern": "^1606-",                                    "manufacturer": "Rockwell Automation / Allen-Bradley","vendor": "Rexel",   "note": "Power supplies",      "priority": 50},

    # ── nVent HOFFMAN / Wiegmann ─────────────────────────────────────────────
    {"prefix": "A##P##",    "pattern": "^A\\d+P\\d+",                            "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Enclosure panels",    "priority": 40},
    {"prefix": "A##N##",    "pattern": "^A\\d+N\\d+",                            "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "NEMA enclosures",     "priority": 40},
    {"prefix": "A##SA##",   "pattern": "^A\\d+SA\\d+",                           "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "NEMA 4X enclosures",  "priority": 40},
    {"prefix": "A##H##",    "pattern": "^A\\d+H\\d+",                            "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Enclosures w/hinges", "priority": 40},
    {"prefix": "DAH",       "pattern": "^DAH\\d+",                               "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Electric heaters",    "priority": 50},
    {"prefix": "EF",        "pattern": "^EF\\d+",                                "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Filter fans",         "priority": 40},
    {"prefix": "TF",        "pattern": "^TF\\d+",                                "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Top fans",            "priority": 40},
    {"prefix": "T1##S",     "pattern": "^T1\\d+S",                               "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Thermostats",         "priority": 40},
    {"prefix": "A-4AXF",    "pattern": "^A-4AXF",                                "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Axial fans",          "priority": 50},
    {"prefix": "WF",        "pattern": "^WF\\d+",                                "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Wiegmann enclosures", "priority": 40},
    {"prefix": "WA",        "pattern": "^WA\\d+",                                "manufacturer": "nVent HOFFMAN",                     "vendor": "Rexel",   "note": "Wiegmann accessories","priority": 40},

    # ── Panduit ──────────────────────────────────────────────────────────────
    {"prefix": "F#X#WH",    "pattern": "^F\\d+X\\d+WH",                          "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct (narrow)",  "priority": 50},
    {"prefix": "H#X#WH",    "pattern": "^H\\d+X\\d+WH",                          "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct (hinged)",  "priority": 50},
    {"prefix": "G#X#WH",    "pattern": "^G\\d+X\\d+WH",                          "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct (wide)",    "priority": 50},
    {"prefix": "E#X#WH",    "pattern": "^E\\d+X\\d+WH",                          "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct",           "priority": 50},
    {"prefix": "C#WH",      "pattern": "^C\\d+WH",                               "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct covers",    "priority": 50},
    {"prefix": "DIN rail",  "pattern": "^(DIN15|DIN35)",                          "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "DIN rails",           "priority": 30},
    {"prefix": "PLT/PLF",   "pattern": "^(PLT|PLF|PLM)\\d+",                     "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Cable ties",          "priority": 40},
    {"prefix": "S#-E",      "pattern": "^S\\d+-E",                               "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Marker cards",        "priority": 40},
    {"prefix": "PCMB-",     "pattern": "^PCMB-",                                 "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Marker books",        "priority": 40},

    # ── Littelfuse / Bussmann / Mersen (Fuses & Protection) ──────────────────
    {"prefix": "LP-CC",     "pattern": "^LP-CC",                                 "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Low-Peak fuses",      "priority": 50},
    {"prefix": "KLDR",      "pattern": "^KLDR",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class CC fuses",      "priority": 50},
    {"prefix": "KLNR",      "pattern": "^KLNR",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class CC fuses",      "priority": 50},
    {"prefix": "FLNR",      "pattern": "^FLNR",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class RK5 fuses",     "priority": 50},
    {"prefix": "FLSR",      "pattern": "^FLSR",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class RK5 fuses",     "priority": 50},
    {"prefix": "JTD-",      "pattern": "^JTD-?\\d+",                             "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class J fuses",       "priority": 50},
    {"prefix": "JLLN",      "pattern": "^JLLN",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class T fuses",       "priority": 50},
    {"prefix": "JLLS",      "pattern": "^JLLS",                                  "manufacturer": "Littelfuse",                        "vendor": "Rexel",   "note": "Class T fuses",       "priority": 50},
    {"prefix": "FNM-",      "pattern": "^FNM-",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Midget fuses",        "priority": 50},
    {"prefix": "FNQ-",      "pattern": "^FNQ-",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Midget fuses",        "priority": 50},
    {"prefix": "FRN-R",     "pattern": "^FRN-?R",                                "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class RK5 fuses",     "priority": 50},
    {"prefix": "FRS-R",     "pattern": "^FRS-?R",                                "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class RK5 fuses",     "priority": 50},
    {"prefix": "LPJ-",      "pattern": "^LPJ-",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class J fuses",       "priority": 50},
    {"prefix": "LPN-RK",    "pattern": "^LPN-RK",                                "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Low-Peak fuses",      "priority": 50},
    {"prefix": "LPS-RK",    "pattern": "^LPS-RK",                                "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Low-Peak fuses",      "priority": 50},
    {"prefix": "NON-",      "pattern": "^NON-",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class K5 fuses",      "priority": 50},
    {"prefix": "NOS-",      "pattern": "^NOS-",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class K5 fuses",      "priority": 50},
    {"prefix": "ATQR",      "pattern": "^ATQR",                                  "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Midget fuses",        "priority": 40},
    {"prefix": "TRS",       "pattern": "^TRS\\d+",                               "manufacturer": "Eaton Bussmann",                    "vendor": "AWC",     "note": "Class RK5 fuses",     "priority": 40},
    {"prefix": "STP",       "pattern": "^STP\\d+",                               "manufacturer": "Mersen",                            "vendor": "Rexel",   "note": "Surge protection",    "priority": 40},
    {"prefix": "STZ",       "pattern": "^STZ\\d+",                               "manufacturer": "Mersen",                            "vendor": "Rexel",   "note": "Surge protection",    "priority": 40},
    {"prefix": "CEP-",      "pattern": "^CEP-",                                  "manufacturer": "Mersen",                            "vendor": "Rexel",   "note": "Fuse holders",        "priority": 40},

    # ── Burndy (Compression Lugs) ────────────────────────────────────────────
    {"prefix": "YA",        "pattern": "^YA\\d+",                                "manufacturer": "Burndy",                            "vendor": "Rexel",   "note": "Compression lugs",    "priority": 50},
    {"prefix": "YAV",       "pattern": "^YAV\\d+",                               "manufacturer": "Burndy",                            "vendor": "Rexel",   "note": "Vibration-resistant",  "priority": 50},
    {"prefix": "KA",        "pattern": "^KA\\d+",                                "manufacturer": "Burndy",                            "vendor": "Rexel",   "note": "Crimping lugs",       "priority": 50},
    {"prefix": "HYLUG",     "pattern": "^HYLUG",                                 "manufacturer": "Burndy",                            "vendor": "Rexel",   "note": "Hylug terminals",     "priority": 50},

    # ── TCI (Line Reactors / Filters) ────────────────────────────────────────
    {"prefix": "KDRH",      "pattern": "^KDRH",                                  "manufacturer": "TCI",                               "vendor": "Rexel",   "note": "Line reactors",       "priority": 50},
    {"prefix": "KDR",       "pattern": "^KDR\\d+",                               "manufacturer": "TCI",                               "vendor": "Rexel",   "note": "Drive reactors",      "priority": 40},
    {"prefix": "KLR",       "pattern": "^KLR\\d+",                               "manufacturer": "TCI",                               "vendor": "Rexel",   "note": "Load reactors",       "priority": 40},
    {"prefix": "HG",        "pattern": "^HG[LP]\\d+",                            "manufacturer": "TCI",                               "vendor": "Rexel",   "note": "Harmonic filters",    "priority": 40},

    # ── Ilsco (Ground Bars / Lugs) ───────────────────────────────────────────
    {"prefix": "2S",        "pattern": "^2S\\d+",                                "manufacturer": "Ilsco",                             "vendor": "Rexel",   "note": "Ground lugs",         "priority": 40},
    {"prefix": "GBL",       "pattern": "^GBL",                                   "manufacturer": "Ilsco",                             "vendor": "Rexel",   "note": "Ground bars",         "priority": 40},
    {"prefix": "PDB",       "pattern": "^PDB",                                   "manufacturer": "Ilsco",                             "vendor": "Rexel",   "note": "Power dist blocks",   "priority": 40},
    {"prefix": "PDBS",      "pattern": "^PDBS",                                  "manufacturer": "Ilsco",                             "vendor": "Rexel",   "note": "Splicer blocks",      "priority": 50},

    # ── Phoenix Contact ──────────────────────────────────────────────────────
    {"prefix": "UT",        "pattern": "^UT\\d+",                                "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Terminal blocks",     "priority": 40},
    {"prefix": "UK",        "pattern": "^UK\\d+",                                "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Terminal blocks",     "priority": 40},
    {"prefix": "PT",        "pattern": "^PT\\s?\\d+",                            "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Push-in terminals",   "priority": 40},
    {"prefix": "DIN-",      "pattern": "^DIN-\\d+",                              "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "DIN terminal blocks", "priority": 40},
    {"prefix": "QUINT",     "pattern": "^QUINT",                                 "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Power supplies",      "priority": 50},
    {"prefix": "TRIO",      "pattern": "^TRIO",                                  "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Power supplies",      "priority": 50},
    {"prefix": "STEP",      "pattern": "^STEP-",                                 "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Power supplies",      "priority": 50},
    {"prefix": "PLC-",      "pattern": "^PLC-",                                  "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Relay modules",       "priority": 40},
    {"prefix": "FL ",       "pattern": "^FL\\s",                                 "manufacturer": "Phoenix Contact",                   "vendor": "AWC",     "note": "Ethernet switches",   "priority": 40},

    # ── Rittal ───────────────────────────────────────────────────────────────
    {"prefix": "WM",        "pattern": "^WM\\d+",                                "manufacturer": "Rittal",                            "vendor": "AWC",     "note": "Enclosures",          "priority": 40},
    {"prefix": "TS",        "pattern": "^TS\\s?\\d{4}",                          "manufacturer": "Rittal",                            "vendor": "AWC",     "note": "TS8 enclosures",      "priority": 40},
    {"prefix": "AE",        "pattern": "^AE\\d{4}",                              "manufacturer": "Rittal",                            "vendor": "AWC",     "note": "AE enclosures",       "priority": 40},

    # ── Hammond Power / Hammond Enclosures ───────────────────────────────────
    {"prefix": "HPS",       "pattern": "^HPS\\w+",                               "manufacturer": "Hammond Power",                     "vendor": "Rexel",   "note": "Transformers",        "priority": 40},
    {"prefix": "EJ",        "pattern": "^EJ\\d{4,}",                             "manufacturer": "Hammond Enclosures",                "vendor": "AWC",     "note": "Enclosures",          "priority": 40},

    # ── Automation Direct ────────────────────────────────────────────────────
    {"prefix": "GS",        "pattern": "^GS[123]-",                              "manufacturer": "Automation Direct",                 "vendor": "Automation Direct", "note": "VFDs (GS series)", "priority": 40},
    {"prefix": "EA",        "pattern": "^EA[27]-",                               "manufacturer": "Automation Direct",                 "vendor": "Automation Direct", "note": "HMI (EA series)",  "priority": 40},
    {"prefix": "BX-DM",     "pattern": "^BX-DM",                                 "manufacturer": "Automation Direct",                 "vendor": "Automation Direct", "note": "BRX PLC",          "priority": 50},
    {"prefix": "P2-",       "pattern": "^P2-",                                   "manufacturer": "Automation Direct",                 "vendor": "Automation Direct", "note": "Productivity PLC", "priority": 40},

    # ── Turck ────────────────────────────────────────────────────────────────
    {"prefix": "NI",        "pattern": "^NI\\d+",                                "manufacturer": "Turck",                             "vendor": "A-Tech",  "note": "Prox sensors",        "priority": 40},
    {"prefix": "BI",        "pattern": "^BI\\d+",                                "manufacturer": "Turck",                             "vendor": "A-Tech",  "note": "Prox sensors",        "priority": 40},
    {"prefix": "TBEN-",     "pattern": "^TBEN-",                                 "manufacturer": "Turck",                             "vendor": "A-Tech",  "note": "Multiprotocol I/O",   "priority": 50},
    {"prefix": "FCS-",      "pattern": "^FCS-",                                  "manufacturer": "Turck",                             "vendor": "A-Tech",  "note": "Flow sensors",        "priority": 40},

    # ── Red Lion ─────────────────────────────────────────────────────────────
    {"prefix": "G0#",       "pattern": "^G0\\d[SCR]\\d+",                        "manufacturer": "Red Lion Controls",                 "vendor": "A-Tech",  "note": "Graphite HMIs",       "priority": 50},
    {"prefix": "G1#",       "pattern": "^G1\\d[SCR]\\d+",                        "manufacturer": "Red Lion Controls",                 "vendor": "A-Tech",  "note": "Graphite HMIs",       "priority": 50},
    {"prefix": "CR1000/3000","pattern": "^CR[13]000",                            "manufacturer": "Red Lion Controls",                 "vendor": "A-Tech",  "note": "CR series HMIs",      "priority": 50},

    # ── Mean Well (Power Supplies) ───────────────────────────────────────────
    {"prefix": "MDR-",      "pattern": "^MDR-",                                  "manufacturer": "Mean Well",                         "vendor": "RS America","note": "DIN rail PSU",       "priority": 40},
    {"prefix": "SDR-",      "pattern": "^SDR-",                                  "manufacturer": "Mean Well",                         "vendor": "RS America","note": "DIN rail PSU",       "priority": 40},
    {"prefix": "NDR-",      "pattern": "^NDR-",                                  "manufacturer": "Mean Well",                         "vendor": "RS America","note": "DIN rail PSU",       "priority": 40},
    {"prefix": "EDR-",      "pattern": "^EDR-",                                  "manufacturer": "Mean Well",                         "vendor": "RS America","note": "DIN rail PSU",       "priority": 40},
    {"prefix": "DRP-",      "pattern": "^DRP-",                                  "manufacturer": "Mean Well",                         "vendor": "RS America","note": "DIN rail PSU",       "priority": 40},

    # ── MOXA ─────────────────────────────────────────────────────────────────
    {"prefix": "EDS-",      "pattern": "^EDS-",                                  "manufacturer": "MOXA",                              "vendor": "AWC",     "note": "Ethernet switches",   "priority": 50},
    {"prefix": "MB31",      "pattern": "^MB31",                                  "manufacturer": "MOXA",                              "vendor": "AWC",     "note": "Serial converters",   "priority": 50},
    {"prefix": "NPort",     "pattern": "^NPORT",                                 "manufacturer": "MOXA",                              "vendor": "AWC",     "note": "Serial servers",      "priority": 50},

    # ── Weidmuller ───────────────────────────────────────────────────────────
    {"prefix": "WDU",       "pattern": "^WDU",                                   "manufacturer": "Weidmuller",                        "vendor": "AWC",     "note": "Terminal blocks",     "priority": 40},
    {"prefix": "WTR",       "pattern": "^WTR",                                   "manufacturer": "Weidmuller",                        "vendor": "AWC",     "note": "Relay modules",       "priority": 40},
    {"prefix": "PRO ",      "pattern": "^PRO\\s",                                "manufacturer": "Weidmuller",                        "vendor": "AWC",     "note": "Power supplies",      "priority": 40},

    # ── Panduit (additional color codes: LG=light gray, BK=black, IW=ivory) ──
    {"prefix": "F#X#LG",    "pattern": "^F\\d+(\\.\\d+)?X\\d+(LG|BK|IW)",       "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct (gray/blk)", "priority": 50},
    {"prefix": "H#X#LG",    "pattern": "^H\\d+(\\.\\d+)?X\\d+(LG|BK|IW)",       "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct hinged",    "priority": 50},
    {"prefix": "G#X#LG",    "pattern": "^G\\d+(\\.\\d+)?X\\d+(LG|BK|IW)",       "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct wide",      "priority": 50},
    {"prefix": "C#LG",      "pattern": "^C\\d+(\\.\\d+)?(LG|BK|IW)",            "manufacturer": "Panduit",                           "vendor": "Rexel",   "note": "Wire duct covers",    "priority": 50},

    # ── Automation Direct / Factorymation ─────────────────────────────────────
    {"prefix": "PD22-",     "pattern": "^PD22-",                                 "manufacturer": "Automation Direct",                 "vendor": "Factorymation", "note": "22mm pilot devices","priority": 50},
    {"prefix": "PB-",       "pattern": "^PB-",                                   "manufacturer": "Automation Direct",                 "vendor": "Factorymation", "note": "Push buttons",     "priority": 40},
    {"prefix": "ZB2-",      "pattern": "^ZB2-",                                  "manufacturer": "Automation Direct",                 "vendor": "Factorymation", "note": "22mm components",  "priority": 40},
    {"prefix": "ZB4-",      "pattern": "^ZB4-",                                  "manufacturer": "Automation Direct",                 "vendor": "Factorymation", "note": "30mm components",  "priority": 40},
]

def _compile_default_prefix_rules():
    """Pre-compile regex patterns for default prefix rules."""
    compiled = []
    for r in _DEFAULT_PREFIX_RULES_RAW:
        try:
            compiled.append({
                "prefix": r["prefix"],
                "pattern": re.compile(r["pattern"], re.IGNORECASE),
                "manufacturer": r["manufacturer"],
                "vendor": r["vendor"],
                "note": r.get("note", ""),
                "priority": r.get("priority", 0),
            })
        except re.error:
            pass
    compiled.sort(key=lambda x: (-x["priority"], -len(x["prefix"])))
    return compiled

_DEFAULT_PREFIX_RULES = _compile_default_prefix_rules()


def load_routing_rules():
    """Load all three tiers of vendor routing rules from DB.
    Returns dict with keys: part_overrides, prefix_rules, vendor_defaults.
    Falls back to hardcoded defaults when DB is unavailable.

    Uses the authenticated user's Supabase client (get_user_sb()) so that
    RLS org-scoped policies work correctly.  Falls back to the global
    anon-key client if called outside a request context.
    """
    rules = {
        "part_overrides": dict(_DEFAULT_PART_OVERRIDES),
        "prefix_rules": list(_DEFAULT_PREFIX_RULES),
        "vendor_defaults": {},
    }

    if not supabase:
        rules["vendor_defaults"] = get_vendor_map()
        return rules

    # Use authenticated client so RLS sees org_id from JWT
    try:
        sb = get_user_sb()
    except Exception:
        sb = supabase  # fallback for non-request contexts

    # Tier 1: Part overrides from DB (merge on top of defaults)
    try:
        rows = sb.table("aae_vendor_part_overrides") \
            .select("part_number,manufacturer,vendor_name,notes") \
            .eq("active", True).execute()
        if rows.data:
            for row in rows.data:
                pn = (row["part_number"] or "").strip().upper()
                if pn:
                    rules["part_overrides"][pn] = {
                        "manufacturer": row.get("manufacturer", ""),
                        "vendor": row["vendor_name"],
                        "note": row.get("notes", ""),
                    }
    except Exception as e:
        print(f"[routing] part_overrides load error: {e}", flush=True)

    # Tier 2: Prefix rules from DB (merge on top of defaults)
    try:
        rows = sb.table("aae_vendor_prefix_rules") \
            .select("prefix,regex_pattern,manufacturer,vendor_name,notes,priority") \
            .eq("active", True) \
            .order("priority", desc=True) \
            .order("prefix") \
            .execute()
        if rows.data:
            db_prefixes = set()
            db_rules = []
            for row in rows.data:
                try:
                    compiled = re.compile(row["regex_pattern"], re.IGNORECASE)
                    db_prefixes.add(row["prefix"].upper())
                    db_rules.append({
                        "prefix": row["prefix"],
                        "pattern": compiled,
                        "manufacturer": row.get("manufacturer", ""),
                        "vendor": row["vendor_name"],
                        "note": row.get("notes", ""),
                        "priority": row.get("priority", 0),
                    })
                except re.error:
                    print(f"[routing] bad regex for prefix '{row['prefix']}': {row['regex_pattern']}", flush=True)
            # Keep defaults that aren't overridden by DB entries
            for dr in _DEFAULT_PREFIX_RULES:
                if dr["prefix"].upper() not in db_prefixes:
                    db_rules.append(dr)
            db_rules.sort(key=lambda x: (-x["priority"], -len(x["prefix"])))
            rules["prefix_rules"] = db_rules
    except Exception as e:
        print(f"[routing] prefix_rules load error: {e}", flush=True)

    # Tier 3: Manufacturer defaults (existing vendor map)
    rules["vendor_defaults"] = get_vendor_map()

    return rules


# ── Manufacturer Alias Normalization ─────────────────────────────────────────
# Maps common manufacturer name variants to the canonical lowercase form used
# in get_vendor_map(). Without this, "Allen-Bradley" → "allen-bradley" would
# miss the "allen bradley" key in vendor_defaults.
_MFR_ALIASES = {
    # Allen-Bradley / Rockwell variants
    "allen-bradley": "allen bradley",
    "a-b": "allen bradley",
    "ab": "allen bradley",
    "rockwell": "allen bradley",
    "rockwell automation": "allen bradley",
    "rockwell automation / allen-bradley": "allen bradley",
    "rockwell/allen-bradley": "allen bradley",
    # Phoenix Contact variants
    "phoenix": "phoenix contact",
    "phx": "phoenix contact",
    # Schneider / Square D variants
    "schneider": "schneider electric",
    "square-d": "square d",
    "sqd": "square d",
    # Eaton / Bussmann variants
    "eaton": "bussmann",
    "eaton bussmann": "bussmann",
    "bussman": "bussmann",
    # Saginaw / SCE variants
    "saginaw control engineering": "saginaw",
    "saginaw control": "saginaw",
    "sce": "saginaw",
    # Automation Direct variants
    "automationdirect": "automation direct",
    "automation-direct": "automation direct",
    "automationdirect.com": "automation direct",
    # Hoffman / nVent variants
    "nvent": "hoffman",
    "nvent hoffman": "hoffman",
    "n-vent": "hoffman",
    # Hammond variants
    "hammond": "hammond power",
    "hammond mfg": "hammond enclosures",
    "hammond manufacturing": "hammond enclosures",
    # Mean Well variants
    "meanwell": "mean well",
    "mean-well": "mean well",
    # Weidmuller variants
    "weidmueller": "weidmuller",
    # Pepperl+Fuchs variants
    "pepperl+fuchs": "pepperl",
    "pepperl fuchs": "pepperl",
    "pepperl-fuchs": "pepperl",
    # LAPP variants
    "lapp group": "lapp",
    "lapp usa": "lapp",
    # Panduit
    "panduit corp": "panduit",
    # Rittal
    "rittal corporation": "rittal",
    # Panduit truncation (OCR sometimes cuts to 5 chars)
    "pandu": "panduit",
    "pandui": "panduit",
    # Saginaw truncation
    "sagin": "saginaw",
    "sagina": "saginaw",
    # Factorymation / Automation Direct
    "automationdirect / factorymation": "factorymation",
    "automation direct / factorymation": "factorymation",
}


def strip_custom_prefix(part_number):
    """Strip Devon/custom prefixes (DVN-, SPC-) to reveal the real manufacturer part number.
    Returns the stripped part number for routing purposes.
    """
    pn = (part_number or "").strip()
    # Strip DVN- prefix (Devon internal numbering)
    if pn.upper().startswith("DVN-"):
        pn = pn[4:]
    # Strip SPC- prefix (Spare Parts Catalog numbering)
    elif pn.upper().startswith("SPC-"):
        pn = pn[4:]
    return pn


def resolve_vendor(part_number, manufacturer, rules=None):
    """Deterministic 4-tier vendor resolution.

    Args:
        part_number:  The part number string (e.g. "1769-OB16")
        manufacturer: The manufacturer string (e.g. "Allen Bradley")
        rules:        Pre-loaded rules from load_routing_rules(). Loads fresh if None.

    Returns:
        dict: {vendor, manufacturer, note, matched_tier}
        matched_tier is one of: "part_override", "prefix_rule", "vendor_default", "UNASSIGNED"
    """
    if rules is None:
        rules = load_routing_rules()

    pn_upper  = (part_number or "").strip().upper()
    mfr_lower = (manufacturer or "").strip().lower()

    # Normalize manufacturer via alias map before Tier 3 lookup
    mfr_lower = _MFR_ALIASES.get(mfr_lower, mfr_lower)

    # Tier 1: Exact part number override (try original first, then stripped)
    if pn_upper and pn_upper in rules["part_overrides"]:
        match = rules["part_overrides"][pn_upper]
        return {
            "vendor": match["vendor"],
            "manufacturer": match.get("manufacturer") or manufacturer,
            "note": match.get("note", ""),
            "matched_tier": "part_override",
        }

    # Strip DVN-/SPC- prefixes for matching
    pn_stripped = strip_custom_prefix(pn_upper).upper()

    # Tier 1b: Check part overrides with stripped part number
    if pn_stripped != pn_upper and pn_stripped in rules["part_overrides"]:
        match = rules["part_overrides"][pn_stripped]
        return {
            "vendor": match["vendor"],
            "manufacturer": match.get("manufacturer") or manufacturer,
            "note": match.get("note", ""),
            "matched_tier": "part_override",
        }

    # Tier 2: Prefix / regex rules (try stripped part number first, then original)
    for pn_try in ([pn_stripped, pn_upper] if pn_stripped != pn_upper else [pn_upper]):
        if pn_try:
            for rule in rules["prefix_rules"]:
                if rule["pattern"].search(pn_try):
                    return {
                        "vendor": rule["vendor"],
                        "manufacturer": rule.get("manufacturer") or manufacturer,
                        "note": rule.get("note", ""),
                        "matched_tier": "prefix_rule",
                    }

    # Tier 3: Manufacturer default
    if mfr_lower and mfr_lower in rules["vendor_defaults"]:
        return {
            "vendor": rules["vendor_defaults"][mfr_lower],
            "manufacturer": manufacturer,
            "note": "",
            "matched_tier": "vendor_default",
        }

    # Tier 4: UNASSIGNED — no guessing
    return {
        "vendor": "UNASSIGNED",
        "manufacturer": manufacturer,
        "note": "",
        "matched_tier": "UNASSIGNED",
    }


def calculate_bid(data):
    r = get_labor_rates()
    enc_qty  = max(1, int(data.get("enc_qty", 1)))
    din_runs = int(data.get("din_rail_runs", 3))
    duct_runs= int(data.get("wire_duct_runs", 4))
    acc_qty  = int(data.get("enc_accessories", 0))
    br_1p    = int(data.get("branch_1p", 0))
    br_2p    = int(data.get("branch_2p", 0))
    br_3p    = int(data.get("branch_3p", 0))
    fused_d  = int(data.get("fused_disconnects", 0))
    cpt      = 1 if str(data.get("cpt_present","N")).upper()=="Y" else 0
    pdb      = int(data.get("pdb_qty", 0))
    main_amp = int(data.get("main_amp", 100))
    relay_ic = int(data.get("relay_icecube", 0))
    relay_dn = int(data.get("relay_din", 0))
    cont_sm  = int(data.get("contactor_small", 0))
    cont_lg  = int(data.get("contactor_large", 0))
    overload = int(data.get("overload", 0))
    timers   = int(data.get("timers", 0))
    ssrs     = int(data.get("ssrs", 0))
    pilots   = int(data.get("pilot_lights", 0))
    selectors= int(data.get("selectors", 0))
    pbs      = int(data.get("push_buttons", 0))
    estops   = int(data.get("estops", 0))
    vfd_sm   = int(data.get("vfd_small", 0))
    vfd_md   = int(data.get("vfd_med", 0))
    vfd_lg   = int(data.get("vfd_large", 0))
    ss_sm    = int(data.get("soft_starter_small", 0))
    ss_lg    = int(data.get("soft_starter_large", 0))
    plc_yn   = str(data.get("plc_present","N")).upper()=="Y"
    di       = int(data.get("plc_di", 0))
    do_pts   = int(data.get("plc_do", 0))
    ai       = int(data.get("plc_ai", 0))
    ao       = int(data.get("plc_ao", 0))
    hmi      = 1 if str(data.get("hmi_present","N")).upper()=="Y" else 0
    safe_r   = 1 if str(data.get("safety_relay","N")).upper()=="Y" else 0
    eth_sw   = 1 if str(data.get("eth_switch","N")).upper()=="Y" else 0
    eth_cab  = int(data.get("eth_cables", 0))
    tb_std   = int(data.get("tb_standard", 0))
    tb_gnd   = int(data.get("tb_ground", 0))
    tb_fsd   = int(data.get("tb_fused", 0))
    tb_dis   = int(data.get("tb_disconnect", 0))
    tb_total = tb_std + tb_gnd + tb_fsd + tb_dis
    markers  = str(data.get("terminal_markers","Y")).upper()=="Y"
    ferrules = str(data.get("ferrules","Y")).upper()=="Y"
    wire_cnt      = int(data.get("wire_count", 0))
    wire_len      = float(data.get("wire_avg_len", 24))
    wire_gauge_key= data.get("wire_gauge", "10_under")
    hs_yn         = str(data.get("heat_shrink","Y")).upper()=="Y"
    fat_hrs       = float(data.get("fat_hours", 0))
    eng_hrs       = float(data.get("eng_hours", 0))
    prog_hrs      = float(data.get("prog_hours", 0))
    comp_key      = data.get("complexity", "STANDARD")
    tech_key      = data.get("tech_level", "JOURNEYMAN")
    expedite      = str(data.get("expedite","N")).upper()=="Y"
    lr_override   = float(data.get("labor_rate_override", 0))
    # Markup/margin: margin_type="margin"|"markup", margin_value=decimal e.g. 0.25
    margin_type   = data.get("margin_type", "markup")
    margin_value  = float(data.get("margin_value", SHOP_RATES["mat_markup"]))
    # Calibration factor from frontend (learned from actual vs estimated hours)
    calib_factor  = float(data.get("calib_factor", 1.0))

    # If wire count not provided, estimate it
    # Revised formula: I/O points drive most wiring on PLC panels
    # tb_total drives field wiring, not all I/O have individual home-run wires
    if wire_cnt == 0:
        wire_cnt = int((di+do_pts)*0.8 + (ai+ao)*1.0 + tb_total*0.6 + (br_2p+br_3p)*2)
        wire_cnt = max(wire_cnt, enc_qty * 10)  # minimum 10 wires per enclosure

    # Labor minutes by section
    enc_min = (r["enclosure_prep"]+r["subpanel_mount"]+r["panel_layout"])*enc_qty + \
              r["din_rail"]*din_runs + r["wire_duct"]*duct_runs*4 + r["enc_accessory"]*acc_qty

    pwr_min = (r["main_breaker_small"] if main_amp<=100 else r["main_breaker_large"])*enc_qty + \
              r["branch_breaker_1p"]*br_1p + r["branch_breaker_23p"]*(br_2p+br_3p) + \
              r["fused_disconnect"]*fused_d + r["cpt"]*cpt + r["pdb"]*pdb

    mc_min  = r["relay_icecube"]*relay_ic + r["relay_din"]*relay_dn + \
              r["contactor_small"]*cont_sm + r["contactor_large"]*cont_lg + \
              r["overload"]*overload + r["timer"]*timers + r["ssr"]*ssrs + \
              r["vfd_small"]*vfd_sm + r["vfd_med"]*vfd_md + r["vfd_large"]*vfd_lg + \
              r["soft_starter_small"]*ss_sm + r["soft_starter_large"]*ss_lg

    cd_min  = r["pilot_light"]*pilots + r["selector"]*selectors + \
              r["pushbutton"]*pbs + r["estop"]*estops

    plc_min = (r["plc_rack"] if plc_yn else 0) + \
              r["plc_di_do"]*(di+do_pts) + r["plc_ai_ao"]*(ai+ao) + \
              r["hmi"]*hmi + r["safety_relay"]*safe_r + \
              r["eth_switch"]*eth_sw + r["eth_cable"]*eth_cab

    tb_min  = r["tb_standard"]*tb_std + r["tb_ground"]*tb_gnd + \
              r["tb_fused"]*tb_fsd + r["tb_disconnect"]*tb_dis + \
              (r["terminal_markers"]*max(1,int(tb_total*1.2/10)) if markers else 0) + \
              r["wire_land_control"]*wire_cnt*2 + \
              (r["ferrule"]*wire_cnt*2 if ferrules else 0) + \
              r["wire_route"]*wire_cnt

    hs_min  = (r["heat_shrink_label"]*wire_cnt*2 +
               r["heat_shrink_batch"]*max(1,int(wire_cnt*2/50)) if hs_yn else 0)

    ul_min  = r["ul_labels"]*enc_qty + r["continuity_check"]*wire_cnt + \
              r["hipot"]*enc_qty + r["as_built"]*enc_qty + r["qc_signoff"]*enc_qty

    raw_min = enc_min+pwr_min+mc_min+cd_min+plc_min+tb_min+hs_min+ul_min
    raw_hrs = raw_min / 60.0

    c_mult    = COMPLEXITY_MULT.get(comp_key, 1.0)
    t_mult    = TECH_MULT.get(tech_key, 1.0)
    # Apply calibration factor to raw hours (learned from actual vs estimated)
    total_hrs = raw_hrs * c_mult * t_mult * calib_factor + fat_hrs + eng_hrs + prog_hrs

    # Wire cost using AAE per-gauge pricing
    wire_cost_per_ft = WIRE_GAUGE_COSTS.get(wire_gauge_key, 0.40)
    ctrl_wire_ft  = int(wire_cnt * wire_len / 12 * 1.15)
    pwr_wire_ft   = int((br_1p + (br_2p+br_3p)*2) * 1.5)
    total_wire_ft = ctrl_wire_ft + pwr_wire_ft
    wire_cost     = total_wire_ft * wire_cost_per_ft

    # Heat shrink: H075X044H1T $275/roll, 1000 labels, 850 usable
    hs_labels_qty  = wire_cnt * 2 if hs_yn else 0
    hs_rolls_needed= max(0, -(-hs_labels_qty // int(SHOP_RATES["heat_shrink_per_roll"])))
    hs_cost        = hs_rolls_needed * SHOP_RATES["heat_shrink_roll_cost"] if hs_yn else 0

    tb_marker_strips = max(0, int(tb_total*1.2/10)) if markers else 0
    ferrule_bags  = max(0, int(wire_cnt*2/500)+1) if ferrules else 0
    din_sticks    = din_runs
    duct_sections = duct_runs * 2
    eth_cables_qty= eth_cab
    spare_tbs     = int(tb_std * 0.1)

    mat_cost = (
        wire_cost +
        hs_cost +
        tb_marker_strips * 2.50 +
        ferrule_bags * 28.00 +
        din_sticks   * 12.50 +
        din_runs*2   * 1.20 +
        duct_sections* 8.75 +
        eth_cables_qty * 8.00 +
        cpt * 12.00 +
        spare_tbs    * 1.10 +
        enc_qty      * 18.00
    )

    # Manual BOM items
    manual_items = data.get("manual_bom", [])
    manual_total = sum(float(i.get("qty",0))*float(i.get("unit_cost",0)) for i in manual_items)
    mat_cost += manual_total

    # Apply markup or margin per user selection
    eff_labor_rate = lr_override if lr_override > 0 else SHOP_RATES["labor_rate"]
    labor_cost = total_hrs * eff_labor_rate
    if margin_type == "margin":
        divisor = max(0.01, 1.0 - margin_value)
        mat_with_markup = mat_cost / divisor
    else:
        mat_with_markup = mat_cost * (1 + margin_value)

    subtotal      = labor_cost + mat_with_markup
    overhead_cost = subtotal * SHOP_RATES["overhead"]
    pre_profit    = subtotal + overhead_cost
    exp_cost      = pre_profit * SHOP_RATES["expedite"] if expedite else 0
    total_price   = pre_profit / (1 - SHOP_RATES["profit"]) + exp_cost

    return {
        "raw_hours": round(raw_hrs, 2),
        "complexity_mult": c_mult,
        "_rates_snapshot": r,   # exact rates used — passed to hours report so line items match
        "tech_mult": t_mult,
        "calib_factor": calib_factor,
        "total_hours": round(total_hrs, 2),
        # Full hour breakdown for transparency report
        "hour_breakdown": {
            "enclosure_hrs":    round(enc_min/60, 2),
            "power_hrs":        round(pwr_min/60, 2),
            "motor_ctrl_hrs":   round(mc_min/60, 2),
            "control_dev_hrs":  round(cd_min/60, 2),
            "plc_network_hrs":  round(plc_min/60, 2),
            "terminals_wire_hrs": round(tb_min/60, 2),
            "heat_shrink_hrs":  round(hs_min/60, 2),
            "ul_qc_hrs":        round(ul_min/60, 2),
            "fat_hrs":          fat_hrs,
            "eng_hrs":          eng_hrs,
            "prog_hrs":         prog_hrs,
            "wire_count_used":  wire_cnt,
            "raw_min_total":    round(raw_min, 1),
        },
        "wire_count": wire_cnt,
        "hs_labels_qty": hs_labels_qty,
        "hs_rolls": hs_rolls_needed,
        "ctrl_wire_ft": ctrl_wire_ft,
        "pwr_wire_ft": pwr_wire_ft,
        "wire_cost": round(wire_cost, 2),
        "hs_cost": round(hs_cost, 2),
        "wire_cost_per_ft": wire_cost_per_ft,
        "mat_cost_raw": round(mat_cost, 2),
        "mat_cost_markup": round(mat_with_markup, 2),
        "margin_type": margin_type,
        "margin_value": margin_value,
        "labor_cost": round(labor_cost, 2),
        "labor_rate_used": eff_labor_rate,
        "subtotal": round(subtotal, 2),
        "overhead_cost": round(overhead_cost, 2),
        "pre_profit": round(pre_profit, 2),
        "expedite_cost": round(exp_cost, 2),
        "total_price": round(total_price, 2),
        "price_per_hour": round(total_price/total_hrs, 2) if total_hrs > 0 else 0,
        "labor_pct": round(labor_cost/total_price*100, 1) if total_price > 0 else 0,
        "mat_pct": round(mat_with_markup/total_price*100, 1) if total_price > 0 else 0,
    }

# ── AI Drawing Scan ────────────────────────────────────────────────────────
_OLD_PROMPT_DEAD_CODE = """placeholder"""
_DEAD_TRIPLE_QUOTE_OPEN = """

═══════════════════════════════════════════════════════════════════
STEP 1 — READ THE BOM TABLE COLUMN HEADERS
═══════════════════════════════════════════════════════════════════
Find the BOM table in the drawing. Read its column headers LEFT TO RIGHT exactly as printed.
Report them in the "column_mapping" field. Then map each header to a data field.

How to identify what each column contains:
  → The PART NUMBER column has codes like "1769-L33ER", "SCE-484812WFLP", "2907719"
     (alphanumeric codes with dashes, slashes, or numbers — NOT plain English descriptions)
  → The MANUFACTURER column has short company names like "ALLEN BRADLEY", "PHOENIX CONTACT",
     "SAGINAW", "ABB", "SIEMENS" (1-3 words, always a company name)
  → The DESCRIPTION column has the LONGEST text — full English sentences describing the part
     like "ENCLOSURE, WALL MOUNT, CARBON STEEL, TYPE 3R/12/13, 48 X 48 X 12"
  → The QTY column has small numbers (1, 2, 5, 32, etc.)
  → The ITEM column has sequential row numbers (1, 2, 3, 4...)

═══════════════════════════════════════════════════════════════════
STEP 2 — EXTRACT ROW BY ROW, COLUMN BY COLUMN
═══════════════════════════════════════════════════════════════════
For each row, go LEFT TO RIGHT across the table, reading each cell under its header.
Place each value into the field that header maps to. Do NOT skip any column.

SELF-CHECK after each row — catch these common mistakes:
  ✗ WRONG: description = "ALLEN BRADLEY" ← That's a company name, not a description!
           If your "description" is just 1-2 words and a company name, you read the MFG column.
  ✓ RIGHT: description = "COMPACTLOGIX, POWER SUPPLY, 24VDC, 120/240VAC INPUT"
           Descriptions are LONG — they describe what the part IS and its specifications.

  ✗ WRONG: manufacturer = "ENCLOSURE, WALL MOUNT, CARBON STEEL, 48 X 48" ← That's a description!
  ✓ RIGHT: manufacturer = "SAGINAW" ← Short company name.

  ✗ WRONG: Same part number repeated 32 times with qty=1 each.
  ✓ RIGHT: One entry with the actual qty from the QTY column (e.g., qty=32).
           Each physical ROW in the BOM table = exactly ONE item in your output.

═══════════════════════════════════════════════════════════════════
STEP 3 — VERIFY PART NUMBERS ARE REAL
═══════════════════════════════════════════════════════════════════
For each part_number in your output, it MUST appear VERBATIM in the PDF.
If you cannot clearly read the characters, write "[UNREADABLE]".
NEVER generate a plausible-looking part number from memory. We will order wrong parts.

Return ONLY valid JSON — no markdown, no explanation. Just the raw JSON object.

{
  "column_mapping": {
    "detected_headers": ["ITEM", "QTY", "CATALOG", "MFG", "DESC"],
    "mapping": {
      "ITEM": "item_num",
      "QTY": "qty",
      "CATALOG": "part_number",
      "MFG": "manufacturer",
      "DESC": "description"
    }
  },
  "extraction_summary": {
    "drawing_types_found": ["BOM", "SCHEMATIC", "TERMINAL_SCHEDULE", "IO_LIST"],
    "confidence": 0.0,
    "scope_gap_flags": [],
    "review_flags": [],
    "total_bom_rows_on_drawing": 0
  },
  "quantities": {
    "enc_qty": 0, "din_rail_runs": 0, "wire_duct_runs": 0, "enc_accessories": 0,
    "main_amp": 0, "main_disconnect_type": "",
    "branch_1p": 0, "branch_2p": 0, "branch_3p": 0,
    "fused_disconnects": 0, "cpt_present": "N", "cpt_kva": 0, "pdb_qty": 0,
    "relay_icecube": 0, "relay_din": 0,
    "contactor_small": 0, "contactor_large": 0, "overload": 0,
    "timers": 0, "ssrs": 0,
    "pilot_lights": 0, "selectors": 0, "push_buttons": 0, "estops": 0,
    "vfd_small": 0, "vfd_med": 0, "vfd_large": 0,
    "soft_starter_small": 0, "soft_starter_large": 0,
    "plc_present": "N", "plc_manufacturer": "", "plc_model": "",
    "plc_di": 0, "plc_do": 0, "plc_ai": 0, "plc_ao": 0,
    "hmi_present": "N", "hmi_size": 0,
    "safety_relay": "N", "eth_switch": "N", "eth_cables": 0,
    "tb_standard": 0, "tb_ground": 0, "tb_fused": 0, "tb_disconnect": 0,
    "wire_count": 0, "wire_avg_len": 24
  },
  "bom_line_items": [
    {
      "item_num": 1,
      "qty": 1,
      "part_number": "SCE-484812WFLP",
      "manufacturer": "SAGINAW",
      "description": "ENCLOSURE, WALL MOUNT, CARBON STEEL, 3PT LATCH, TYPE 3R/12/13, 48 X 48 X 12",
      "unit": "ea",
      "category": "Enclosure",
      "notes": ""
    }
  ]
}

Rules:
- VFDs: <=5HP = small, 6-25HP = med, 26-100HP = large
- Contactors: <=40A = small, >40A = large
- Soft starters: <=50A = small, >50A = large
- Count wire numbers if a wire schedule exists for wire_count
- If no wire schedule: estimate wire_count as (DI+DO)*0.8 + (AI+AO)*1.0 + terminals*0.6
- bom_line_items: extract EVERY SINGLE line item from any BOM table found in the drawings
  - THIS IS THE MOST IMPORTANT PART OF YOUR JOB. The BOM table is the primary deliverable.
  - Do NOT skip or omit ANY rows. Each physical row in the BOM table = exactly ONE item.
  - If the BOM table has 48 rows, you MUST return exactly 48 items. Count them carefully.
  - ONE ROW = ONE ITEM. If a row says QTY 32, output qty:32 — do NOT create 32 separate items.
  - The number of items in bom_line_items MUST equal total_bom_rows_on_drawing.
  - If a BOM table has no manufacturer column, leave manufacturer as "" — do NOT guess.
  - If a BOM table has no description column, leave description as "" and note it in notes.
  - Use these categories: Enclosure, Power, Motor Ctrl, Control Devices, PLC/Network, Terminals, Relays, Wiring, HMI/Computer, Markers, Other
  - If no BOM table found, return empty array []
  - Set qty to numeric value (not "A/R" — use 1 for A/R items and note "A/R" in notes field)
  - Set total_bom_rows_on_drawing to how many rows you count in the BOM table
- Flag anything uncertain or hard to read in review_flags
- confidence: 0.0 to 1.0"""
# (old single-prompt scan_drawing removed -- see scanner_pipeline.py for v2 multi-stage pipeline)

def scan_drawing(pdf_b64, filename="drawing.pdf"):
    """Thin wrapper that delegates to the multi-stage pipeline in scanner_pipeline.py.
    Passes the module-level claude_client so the pipeline can make API calls."""
    return _pipeline_scan_drawing(claude_client, pdf_b64, filename)


# ── BOM Converter — AI Parser for QuickBooks BOMs & Vendor Quotes ─────────
def convert_bom_pdf(pdf_b64, filename="bom.pdf"):
    """Parse a QuickBooks BOM report or vendor quote PDF.
    Extracts panel names (from 'Item Name/Number' headers) and all line items.
    Returns structured JSON with panels[], each containing line_items[].
    """
    prompt = """You are an expert at parsing QuickBooks Bill of Materials (BOM) reports and vendor quotes for an industrial control panel shop (AAE Automation, UL-508A certified).

Analyze this PDF and extract ALL panels/BOMs and their line items.

═══════════════════════════════════════════════════════════════════
STEP 1 — READ COLUMN HEADERS FIRST, THEN MAP THEM
═══════════════════════════════════════════════════════════════════
Before extracting data, read each BOM table's column headers exactly as printed.
Different documents use different column names. YOU must figure out which column is which.

Map each column header to a data field:
  → part_number:  the catalog/part/model number (could be: CATALOG, CAT NO, PART NUMBER,
                  P/N, ITEM NAME/NUM, MODEL, MFR PART NO, ORDER NUMBER, etc.)
  → manufacturer: the maker/brand (could be: MFG, MANUFACTURER, MFR, VENDOR, BRAND, MAKE, etc.)
  → description:  what the part is (could be: DESC, DESCRIPTION, ITEM DESCRIPTION, REMARKS, etc.)
  → qty:          how many (could be: QTY, QUANTITY, COUNT, EA, AMT, etc.)
  → unit:         unit of measure (could be: U/M, UOM, UNIT, etc.)
  → cost:         per-unit price (could be: COST, PRICE, UNIT COST, UNIT PRICE, RATE, etc.)

CRITICAL RULES:
  1. NEVER INVENT OR HALLUCINATE DATA. Copy text CHARACTER BY CHARACTER from the PDF.
     If you cannot read a cell, write "[UNREADABLE]". A wrong part number is worse than none.
  2. DO NOT MIX UP COLUMNS. Use your header mapping. Manufacturer goes in "manufacturer",
     description goes in "description". Do NOT swap them.
  3. COPY EXACTLY — do not paraphrase, abbreviate, or "fix" any value.

Return ONLY valid JSON — no markdown, no explanation. Just the raw JSON object.

{
  "panels": [
    {
      "panel_name": "DVN-100HP-VFD",
      "document_type": "quickbooks_bom",
      "purchase_description": "Devon 100HP VFD, Rev. 1",
      "column_mapping": {
        "detected_headers": ["Item Name/Num", "Description", "Type", "Cost", "Qty", "U/M", "Total"],
        "mapping": {"Item Name/Num": "part_number", "Description": "description", "Qty": "qty", "U/M": "unit", "Cost": "cost"}
      },
      "line_items": [
        {
          "item_num": 1,
          "part_number": "WF100LP",
          "description": "Type 3R Drive Enclosure with (2) 10 inch Fans",
          "qty": 1,
          "unit": "ea",
          "cost": 1947.97,
          "manufacturer": "",
          "category": "Enclosure",
          "notes": ""
        }
      ]
    }
  ],
  "extraction_summary": {
    "total_panels_found": 1,
    "total_line_items": 4,
    "confidence": 0.95,
    "review_flags": []
  }
}

Rules:
- A QuickBooks BOM starts with a header section containing "Item Name/Number" — use the value after this as panel_name.
  Example: "Item Name/Number  DVN-100HP-VFD" → panel_name = "DVN-100HP-VFD"
  Set document_type = "quickbooks_bom"
- The Purchase Description field below it (e.g., "Devon 100HP VFD, Rev. 1") goes into purchase_description
- If no "Item Name/Number" header is found (vendor quote / price list), set panel_name to "Unknown Panel"
  and document_type = "quote". If there are multiple such quotes, append a number: "Unknown Panel 2", "Unknown Panel 3"
- A single PDF may contain MANY BOMs back-to-back (each starts with its own "Item Name/Number" header section). Extract ALL of them.
- THIS IS CRITICAL: Extract EVERY SINGLE line item from EVERY BOM table. Do NOT skip or omit ANY rows.
  If a BOM has 30 rows, return exactly 30 items. Count them.
- Read the ACTUAL column headers for each table — do NOT assume column positions.
  Use your column_mapping to place each cell's value in the correct output field.
- MANUFACTURER IDENTIFICATION: Even if the document has no manufacturer column, you MUST identify the
  manufacturer from the part number prefix/pattern and description. Use this reference:

  ALLEN-BRADLEY / ROCKWELL: 1756-, 1769-, 1734-, 1794-, 5094-, 5069-, 5034-, 2080-, 1783-, 1492-,
    1497-, 1606-, 20F, 20G, 20P, 22B, 22C, 25A, 25B, 20-HIM, 20-750, 20-COMM, 140G-, 140M-,
    150-, 100-, 100S-, 193-, 199-, 509-, 700-, 700S-, 800F, 800T, 800H, 440R-, 440G-, 2711R-, 2711P-, 6189-
  nVent HOFFMAN: A##P## (panels), A##N## (enclosures), DAH (heaters), EF/TF (fans), WF (Wiegmann), WA
  PANDUIT: F#X#WH, H#X#WH, G#X#WH, C#WH (wire duct/covers), PLT/PLF (ties), S#-E (markers), PCMB-
  LITTELFUSE: LP-CC, KLDR, KLNR, FLNR, FLSR, JTD, JLLN, JLLS
  EATON BUSSMANN: FNM-, FNQ-, FRN-R, FRS-R, LPJ-, LPN-RK, LPS-RK, NON-, NOS-, ATQR, TRS
  MERSEN: STP, STZ, CEP-
  BURNDY: YA## (lugs), YAV, KA##, HYLUG
  TCI: KDRH, KDR, KLR, HGL/HGP (reactors/filters)
  ILSCO: 2S (ground lugs), GBL (ground bars), PDB/PDBS (power dist blocks)
  PHOENIX CONTACT: UT, UK, PT (terminals), QUINT/TRIO/STEP (PSU), PLC- (relays), FL (switches)
  SCHNEIDER ELECTRIC: QO, 9070TF, LV, HOM, NQ, BGA/BJA/HJA/JJA
  RITTAL: WM, TS, AE (enclosures)
  MEAN WELL: MDR-, SDR-, NDR-, EDR-, DRP-
  AUTOMATION DIRECT: GS#-, EA#-, BX-DM, P2-
  TURCK: NI##, BI## (sensors), TBEN- (I/O), FCS-
  MOXA: EDS-, MB31, NPort
  HAMMOND POWER: HPS (transformers)

  IMPORTANT: Devon part numbers often have a "DVN-" prefix. Strip it to identify the real part.
  Example: "DVN-1492-JG4" → real part is "1492-JG4" → Allen-Bradley terminal block.
  Similarly, "SPC-" is a spare parts prefix — strip it too.

  If you still cannot identify the manufacturer, leave it as empty string — do NOT guess.
- Assign each line item a category from: Enclosure, Power, Motor Ctrl, Control Devices, PLC/Network,
  Terminals, Relays, HMI/Computer, Wiring, Markers, Other
  Use "Other" if the category is unclear.
- Set qty to numeric value. If "A/R" use 1 and put "A/R" in notes field.
- Set total_panels_found and total_line_items accurately in extraction_summary
- Flag anything uncertain in review_flags
- confidence: 0.0 to 1.0"""

    import time

    models_to_try = ["claude-sonnet-4-20250514", "claude-haiku-4-5-20251001"]

    for attempt, model in enumerate(models_to_try):
        try:
            api_kwargs = {
                "model": model,
                "max_tokens": 40000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_b64
                            }
                        },
                        {"type": "text", "text": prompt}
                    ]
                }]
            }

            if "sonnet" in model:
                api_kwargs["thinking"] = {"type": "enabled", "budget_tokens": 16000}
                print(f"BOM_CONVERT: Using extended thinking with {model} (16K thinking budget)", flush=True)
            else:
                api_kwargs["temperature"] = 0

            with claude_client.messages.stream(**api_kwargs) as stream:
                response = stream.get_final_message()

            stop_reason = response.stop_reason
            tokens_used = response.usage.output_tokens if response.usage else 0
            if stop_reason == "max_tokens":
                print(f"BOM_CONVERT WARNING: Response truncated at {tokens_used} output tokens (max_tokens hit). Some panels may be incomplete.", flush=True)

            # Extract the text block (skip thinking blocks)
            raw = ""
            for block in response.content:
                if block.type == "text":
                    raw = block.text.strip()
                    break
            if not raw:
                raw = response.content[0].text.strip()
            raw = re.sub(r"^```json\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)

            try:
                result = json.loads(raw)
            except json.JSONDecodeError:
                print(f"BOM_CONVERT: JSON truncated, attempting repair (len={len(raw)})", flush=True)
                last_brace = raw.rfind('}')
                if last_brace > 0:
                    trimmed = raw[:last_brace+1]
                    opens_b  = trimmed.count('{') - trimmed.count('}')
                    opens_sq = trimmed.count('[') - trimmed.count(']')
                    trimmed += ']' * opens_sq + '}' * opens_b
                    try:
                        result = json.loads(trimmed)
                        result["_truncated"] = True
                        print("BOM_CONVERT: JSON repair succeeded", flush=True)
                    except json.JSONDecodeError as e2:
                        raise json.JSONDecodeError(
                            f"JSON parse failed after repair attempt: {e2.msg} (original len={len(raw)})",
                            e2.doc, e2.pos
                        )
                else:
                    raise

            result["_model_used"] = model
            result["_stop_reason"] = stop_reason
            result["_output_tokens"] = tokens_used
            if stop_reason == "max_tokens":
                result["_truncated"] = True
                if "extraction_summary" not in result:
                    result["extraction_summary"] = {}
                if "review_flags" not in result.get("extraction_summary", {}):
                    result["extraction_summary"]["review_flags"] = []
                result["extraction_summary"]["review_flags"].append(
                    "WARNING: Response was truncated — some panels or line items may be missing. Try uploading fewer BOMs per file."
                )
            return result

        except Exception as e:
            err_str = str(e)
            print(f"BOM_CONVERT attempt {attempt+1} ({model}) ERROR: {err_str}")

            if "429" in err_str or "rate_limit" in err_str.lower() or "overloaded" in err_str.lower():
                is_token_limit = "input tokens per minute" in err_str.lower()

                if is_token_limit:
                    print(f"BOM_CONVERT: PDF exceeds token rate limit — skipping Haiku fallback", flush=True)
                    return {
                        "error": "token_rate_limit",
                        "error_message": "This PDF is too large for your current Anthropic API tier. "
                                         "Go to console.anthropic.com → Billing → load more credit to increase your tier, "
                                         "or upload fewer BOMs per file.",
                        "panels": [],
                        "extraction_summary": {"confidence": 0, "total_panels_found": 0, "total_line_items": 0,
                                               "review_flags": ["PDF exceeds API tier token limit."]}
                    }

                if attempt == 0:
                    print(f"BOM_CONVERT: Rate limit on {model}, waiting 30s then retrying...", flush=True)
                    time.sleep(30)
                    models_to_try.insert(attempt + 1, model)
                    continue
                elif attempt < len(models_to_try) - 1:
                    time.sleep(2)
                    continue
                return {
                    "error": "rate_limit",
                    "error_message": "API rate limit reached. Please wait 60 seconds and try again.",
                    "panels": [],
                    "extraction_summary": {"confidence": 0, "total_panels_found": 0, "total_line_items": 0,
                                           "review_flags": ["rate_limit"]}
                }

            import traceback
            err_detail = traceback.format_exc()
            print("BOM_CONVERT ERROR DETAIL:", err_detail)
            return {"error": str(e), "error_detail": err_detail,
                    "panels": [],
                    "extraction_summary": {"confidence": 0, "total_panels_found": 0, "total_line_items": 0}}

    return {"error": "all_models_failed", "panels": [],
            "extraction_summary": {"confidence": 0, "total_panels_found": 0, "total_line_items": 0}}


def enrich_manufacturers(panels, routing_rules=None):
    """Post-extraction manufacturer enrichment.

    Two-pass approach:
      Pass 1 (deterministic): Run every part number through the prefix rules engine.
              This catches everything the regex patterns cover with zero API cost.
      Pass 2 (AI): For remaining unknowns, batch them into a single cheap AI call
              that identifies manufacturers from part numbers + descriptions.

    Modifies panels in-place and returns the count of enriched items.
    """
    if routing_rules is None:
        routing_rules = load_routing_rules()

    enriched_count = 0
    unknown_items = []  # items still missing manufacturer after Pass 1

    # ── Pass 1: Deterministic prefix/override matching ───────────────────
    for panel in panels:
        for item in panel.get("line_items", []):
            if item.get("manufacturer"):
                continue  # already has manufacturer

            pn = (item.get("part_number") or "").strip()
            if not pn:
                continue

            # Try resolve_vendor which now handles DVN-/SPC- stripping
            result = resolve_vendor(pn, "", routing_rules)
            if result["matched_tier"] in ("part_override", "prefix_rule"):
                item["manufacturer"] = result["manufacturer"]
                enriched_count += 1
            else:
                unknown_items.append(item)

    # ── Pass 2: AI enrichment for remaining unknowns ─────────────────────
    if not unknown_items or not claude_client:
        return enriched_count

    # Build a compact list for the AI
    unknown_list = []
    for item in unknown_items:
        pn = (item.get("part_number") or "").strip()
        desc = (item.get("description") or "").strip()
        if pn:
            unknown_list.append({"pn": pn, "desc": desc})

    if not unknown_list:
        return enriched_count

    # Limit batch size to avoid excessive cost (typically < 50 unknowns)
    if len(unknown_list) > 100:
        unknown_list = unknown_list[:100]

    prompt = f"""You are an expert at identifying industrial electrical component manufacturers from part numbers.

For each part number below, identify the manufacturer. If a part starts with "DVN-" or "SPC-", strip that prefix first.

Common manufacturers in this industry:
- Allen-Bradley/Rockwell: 1756-, 1769-, 1734-, 1794-, 5094-, 1492-, 1497-, 20F, 140G-, 150-, 100-, 700-, 800F/T/H, 199-, 2711-, etc.
- nVent HOFFMAN: A##P##, A##N##, DAH, EF, TF, WF, WA
- Panduit: F#X#WH, C#WH, H#X#WH, PLT, S#-E, PCMB-
- Littelfuse: LP-CC, KLDR, KLNR, FLNR, FLSR, JTD, JLLN
- Eaton Bussmann: FNM-, FNQ-, FRN-R, FRS-R, LPJ-, TRS, ATQR, NON-, NOS-
- Mersen: STP, STZ, CEP-
- Burndy: YA##, YAV, KA##, HYLUG
- TCI: KDRH, KDR, KLR
- Ilsco: 2S, GBL, PDB, PDBS
- Phoenix Contact: UT, UK, PT, QUINT, TRIO, STEP-, PLC-, FL
- Schneider Electric: QO, 9070TF, LV, HOM
- Rittal: WM, TS, AE
- Hammond Power: HPS
- Mean Well: MDR-, SDR-, NDR-, DRP-
- Automation Direct: GS#-, EA#-, BX-DM, P2-
- Turck: NI##, BI##, TBEN-, FCS-
- MOXA: EDS-, MB31, NPort
- 3M: 5-8 digit numeric part numbers for electrical products
- Essex Wire/Encore Wire: wire products starting with "EG WI"
- Generic hardware (Hillman, etc.): OLW, OHN bolt/screw part numbers

Return ONLY valid JSON — a list of objects with "pn" and "manufacturer" fields.
If you cannot identify a manufacturer, use empty string "".

Input parts:
{json.dumps(unknown_list, indent=2)}"""

    try:
        response = claude_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4000,
            temperature=0,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        ai_results = json.loads(raw)

        # Build lookup: part_number -> manufacturer
        ai_mfr_map = {}
        for r in ai_results:
            pn = (r.get("pn") or "").strip()
            mfr = (r.get("manufacturer") or "").strip()
            if pn and mfr:
                ai_mfr_map[pn.upper()] = mfr

        # Apply AI results to unknown items
        for item in unknown_items:
            pn = (item.get("part_number") or "").strip().upper()
            if pn in ai_mfr_map:
                item["manufacturer"] = ai_mfr_map[pn]
                enriched_count += 1

        print(f"MFR_ENRICH: AI identified {len(ai_mfr_map)} manufacturers from {len(unknown_list)} unknowns", flush=True)

    except Exception as e:
        print(f"MFR_ENRICH: AI enrichment failed (non-fatal): {e}", flush=True)

    return enriched_count


# ── PDF Quote Generator ────────────────────────────────────────────────────
def generate_quote_pdf(bid_data, calc_results, quote_number):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    navy = colors.HexColor("#1F3864")
    blue = colors.HexColor("#2E74B5")
    orange = colors.HexColor("#C55A11")
    light_gray = colors.HexColor("#F2F2F2")
    pale_blue = colors.HexColor("#DEEAF1")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=20, textColor=colors.white,
                                 fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle("sub", fontSize=10, textColor=colors.HexColor("#BDD7EE"),
                                 fontName="Helvetica", alignment=TA_CENTER)
    section_style=ParagraphStyle("section", fontSize=11, textColor=colors.white,
                                 fontName="Helvetica-Bold", alignment=TA_LEFT)
    label_style = ParagraphStyle("label", fontSize=10, textColor=navy,
                                 fontName="Helvetica-Bold")
    value_style = ParagraphStyle("value", fontSize=10, textColor=colors.black,
                                 fontName="Helvetica")
    total_style = ParagraphStyle("total", fontSize=14, textColor=colors.white,
                                 fontName="Helvetica-Bold", alignment=TA_RIGHT)
    note_style  = ParagraphStyle("note", fontSize=8, textColor=colors.HexColor("#595959"),
                                 fontName="Helvetica-Oblique")

    story = []

    # Header
    header_data = [[Paragraph("AAE AUTOMATION", title_style)],
                   [Paragraph("UL-NNNY  |  UL-508A Certified  |  Industrial Control Panel Specialists", sub_style)]]
    header_table = Table(header_data, colWidths=[7*inch])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), navy),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [navy]),
        ("TOPPADDING", (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 16),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.2*inch))

    # Quote info row
    info_data = [
        [Paragraph("<b>QUOTATION</b>", ParagraphStyle("q", fontSize=16, textColor=orange, fontName="Helvetica-Bold")),
         Paragraph(f"<b>Quote #:</b> {quote_number}<br/>"
                   f"<b>Date:</b> {datetime.now().strftime('%B %d, %Y')}<br/>"
                   f"<b>Valid For:</b> 30 Days",
                   ParagraphStyle("qi", fontSize=10, fontName="Helvetica", leading=16))],
    ]
    info_table = Table(info_data, colWidths=[3.5*inch, 3.5*inch])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("TOPPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.1*inch))
    story.append(HRFlowable(width="100%", thickness=2, color=navy))
    story.append(Spacer(1, 0.1*inch))

    # Customer / Project info
    customer = bid_data.get("customer_name", "")
    project  = bid_data.get("project_name", "")
    estimator= bid_data.get("estimator_name", "")
    if customer or project:
        proj_data = [
            ["Customer:", customer, "Project / PO:", project],
            ["Estimator:", estimator, "Panel Type:", bid_data.get("complexity","STANDARD")],
        ]
        proj_table = Table(proj_data, colWidths=[1.2*inch, 2.3*inch, 1.2*inch, 2.3*inch])
        proj_table.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0,0), (0,-1), navy),
            ("TEXTCOLOR", (2,0), (2,-1), navy),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(proj_table)
        story.append(Spacer(1, 0.15*inch))

    # Cost Summary section
    def section_hdr(text):
        t = Table([[Paragraph(text, section_style)]], colWidths=[7*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), blue),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING", (0,0), (-1,-1), 10),
        ]))
        return t

    story.append(section_hdr("COST SUMMARY"))
    story.append(Spacer(1, 0.05*inch))

    c = calc_results
    cost_rows = [
        ["Total Billable Hours", f"{c['total_hours']:,.2f} hrs",
         f"({c['raw_hours']:,.2f} raw × {c['complexity_mult']}× complexity × {c['tech_mult']}× tech)"],
        ["Labor Rate", f"${c['labor_rate_used']:,.2f}/hr", ""],
        ["Labor Cost", f"${c['labor_cost']:,.2f}", ""],
        ["Material Cost (w/ markup)", f"${c['mat_cost_markup']:,.2f}", f"Raw: ${c['mat_cost_raw']:,.2f}"],
        ["Overhead", f"${c['overhead_cost']:,.2f}", f"{int(SHOP_RATES['overhead']*100)}%"],
    ]
    if c['expedite_cost'] > 0:
        cost_rows.append(["Expedite Surcharge", f"${c['expedite_cost']:,.2f}", "15%"])

    cost_table = Table(cost_rows, colWidths=[2.5*inch, 1.8*inch, 2.7*inch])
    cost_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1), navy),
        ("TEXTCOLOR", (2,0), (2,-1), colors.HexColor("#595959")),
        ("FONTSIZE", (2,0), (2,-1), 9),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, light_gray]),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("LINEBELOW", (0,-1), (-1,-1), 1, navy),
    ]))
    story.append(cost_table)
    story.append(Spacer(1, 0.05*inch))

    # Total price box
    total_data = [[Paragraph(f"TOTAL QUOTED PRICE:  ${c['total_price']:,.2f}", total_style)]]
    total_table = Table(total_data, colWidths=[7*inch])
    total_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), navy),
        ("TOPPADDING", (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 16),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 0.2*inch))

    # Sanity checks
    story.append(section_hdr("ESTIMATING METRICS"))
    story.append(Spacer(1, 0.05*inch))
    metrics = [
        ["Price Per Billable Hour", f"${c['price_per_hour']:,.2f}",
         "Labor %", f"{c['labor_pct']}%"],
        ["Heat Shrink Labels", f"{c['hs_labels_qty']:,} labels",
         "Material %", f"{c['mat_pct']}%"],
        ["Control Wire", f"{c['ctrl_wire_ft']:,} ft",
         "Power Wire", f"{c['pwr_wire_ft']:,} ft"],
        ["Est. Wire Count", f"{c['wire_count']:,} wires",
         "Wire Cost", f"${c['ctrl_wire_ft']*SHOP_RATES['wire_16_per_ft'] + c['pwr_wire_ft']*SHOP_RATES['wire_12_per_ft']:,.2f}"],
    ]
    m_table = Table(metrics, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
    m_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1), navy),
        ("TEXTCOLOR", (2,0), (2,-1), navy),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, light_gray]),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(m_table)
    story.append(Spacer(1, 0.2*inch))

    # Scope / notes
    story.append(section_hdr("SCOPE & TERMS"))
    story.append(Spacer(1, 0.05*inch))
    scope_items = [
        "Panel design and build per UL-508A standards",
        "UL labeling and NNNY certification included" if str(bid_data.get("ul_required","Y")).upper()=="Y" else "",
        "Engineering/design included" if str(bid_data.get("engineering","N")).upper()=="Y" else "Engineering/design NOT included",
        "Programming included" if str(bid_data.get("programming","N")).upper()=="Y" else "Programming NOT included",
        "Witness test included" if str(bid_data.get("witness_test","N")).upper()=="Y" else "",
        "Shipping NOT included — FOB our shop",
        "Quote valid for 30 days from date above",
        "Lead time TBD based on schedule at time of order",
    ]
    for item in scope_items:
        if item:
            story.append(Paragraph(f"• {item}", ParagraphStyle("scope", fontSize=10,
                         fontName="Helvetica", leftIndent=12, spaceAfter=3)))
    story.append(Spacer(1, 0.3*inch))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#BFBFBF")))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(
        "AAE Automation  |  UL-NNNY  |  UL-508A  |  This quotation is confidential and intended solely for the addressee.",
        ParagraphStyle("footer", fontSize=8, textColor=colors.HexColor("#595959"),
                       fontName="Helvetica", alignment=TA_CENTER)))

    doc.build(story)
    buffer.seek(0)
    return buffer

# ── Routes ─────────────────────────────────────────────────────────────────
@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory("static", filename)


@app.route("/api/test")
@require_auth
def api_test():
    """Quick health check — verifies Anthropic API key and model are working."""
    try:
        resp = claude_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=20,
            messages=[{"role":"user","content":"Reply with the word OK only."}]
        )
        return jsonify({"status": "ok", "reply": resp.content[0].text,
                        "anthropic_sdk": "ok"})
    except Exception as e:
        import traceback
        return jsonify({"status": "error", "error": str(e),
                        "detail": traceback.format_exc()}), 500


@app.route("/")
def index():
    return render_template(
        "index.html",
        supabase_url=os.environ.get("SUPABASE_URL", ""),
        supabase_anon_key=os.environ.get("SUPABASE_ANON_KEY", ""),
    )

@app.route("/api/me", methods=["GET"])
@require_auth
def me():
    """Return current user's profile (role, display name). Used by frontend for role-gated UI."""
    # Role is authoritative from JWT app_metadata (set by provision_admin.py).
    # profiles table is supplementary — display name only.
    # If the row doesn't exist yet (user not yet provisioned), we fall back to
    # JWT values so login still works and the admin can see who is unprovisioned.
    user_role = g.user.get("role", "viewer")
    user_email = g.user["email"]
    display_name = ""

    if _sb_url:
        try:
            sb = get_user_sb()
            result = sb.table("profiles").select("display_name,username").eq("user_id", g.user["id"]).maybe_single().execute()
            if result.data:
                display_name = result.data.get("display_name", "") or result.data.get("username", "")
        except Exception:
            pass  # profiles row missing or RLS blocked — not fatal, role comes from JWT

    return jsonify({
        "id":           g.user["id"],
        "email":        user_email,
        "aal":          g.user.get("aal"),
        "role":         user_role,
        "display_name": display_name or user_email,
        "org_id":       g.user.get("org_id"),
    })

MAX_SCAN_SIZE = 25 * 1024 * 1024  # 25 MB
PDF_MAGIC = b"%PDF"

@app.route("/api/scan", methods=["POST"])
@require_auth
def scan():
    print("=== /api/scan called ===", flush=True)
    if "drawing" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["drawing"]

    # Read once; enforce size
    pdf_bytes = f.read()
    if len(pdf_bytes) > MAX_SCAN_SIZE:
        return jsonify({"error": f"File too large (max 25 MB)"}), 413

    # Enforce PDF by magic bytes
    if not pdf_bytes.startswith(PDF_MAGIC):
        return jsonify({"error": "Only PDF files are accepted"}), 415

    # Audit the upload
    file_hash = hashlib.sha256(pdf_bytes).hexdigest()[:16]
    audit_log("scan_upload", "drawing", file_hash, {
        "filename": f.filename,
        "size_bytes": len(pdf_bytes),
    })

    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    print(f"API key present: {bool(api_key)}", flush=True)
    result = scan_drawing(pdf_b64, f.filename)
    if "error" in result:
        print(f"SCAN ERROR: {result['error']}", flush=True)
    return jsonify(result)

@app.route("/api/calculate", methods=["POST"])
@require_auth
def calculate():
    import traceback
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data received"}), 400
        result = calculate_bid(data)
        return jsonify(result)
    except Exception as e:
        err = traceback.format_exc()
        print("CALCULATE ERROR:", err, flush=True)
        return jsonify({"error": str(e), "detail": err}), 500

@app.route("/api/save_bid", methods=["POST"])
@require_auth
def save_bid():
    data = request.get_json()
    try:
        row = {
            "customer_name":  data.get("customer_name",""),
            "project_name":   data.get("project_name",""),
            "estimator_name": data.get("estimator_name",""),
            "complexity":     data.get("complexity","STANDARD"),
            "tech_level":     data.get("tech_level","JOURNEYMAN"),
            "total_hours":    data.get("calc",{}).get("total_hours",0),
            "total_price":    data.get("calc",{}).get("total_price",0),
            "bid_data":       json.dumps(data),
            "created_by":     g.user["id"],
            "created_at":     datetime.now().isoformat(),
        }
        if not _sb_url: return jsonify({"error": "Supabase not configured"}), 500
        sb = get_user_sb()
        result = sb.table("bids").insert(row).execute()
        bid_id = result.data[0]["id"]
        audit_log("save_bid", "bid", bid_id, {
            "customer": row["customer_name"],
            "project": row["project_name"],
            "total_price": row["total_price"],
        })
        return jsonify({"success": True, "id": bid_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/bids", methods=["GET"])
@require_auth
def get_bids():
    try:
        if not _sb_url: return jsonify([])
        sb = get_user_sb()
        result = sb.table("bids").select("*").is_("is_deleted", "null").order("created_at", desc=True).limit(50).execute()
        return jsonify(result.data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/bids/<int:bid_id>", methods=["DELETE"])
@require_auth
def delete_bid(bid_id):
    """Soft delete a bid — admin only. Sets is_deleted = now() instead of removing the row."""
    if g.user.get("role") != "admin":
        return jsonify({"error": "Admin role required"}), 403
    try:
        sb = get_user_sb()
        from datetime import timezone
        sb.table("bids").update({
            "is_deleted": datetime.now(timezone.utc).isoformat()
        }).eq("id", bid_id).execute()
        audit_log("bid_soft_delete", "bids", str(bid_id), {"deleted_by": g.user["email"]})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/quote_pdf", methods=["POST"])
@require_auth
def quote_pdf():
    data = request.get_json()
    bid_data  = data.get("bid_data", {})
    calc_data = data.get("calc", {})
    quote_num = f"AAE-{datetime.now().strftime('%Y%m%d')}-{str(datetime.now().microsecond)[:4]}"
    pdf_buf   = generate_quote_pdf(bid_data, calc_data, quote_num)
    return send_file(pdf_buf, mimetype="application/pdf",
                     as_attachment=True,
                     download_name=f"AAE_Quote_{quote_num}.pdf")

@app.route("/api/bom_pdf", methods=["POST"])
@require_auth
def bom_pdf():
    """Generate a BOM PDF for download."""
    data      = request.get_json()
    bid_data  = data.get("bid_data", {})
    calc_data = data.get("calc", {})
    quote_num = data.get("quote_num", f"AAE-{datetime.now().strftime('%Y%m%d')}")
    buf = generate_bom_pdf(bid_data, calc_data, quote_num)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"AAE_BOM_{quote_num}.pdf")


@app.route("/api/bom_excel", methods=["POST"])
@require_auth
def bom_excel():
    """Generate post-calculation BOM Excel using the same premium format as bom_from_scan.
    Builds BOM from calc data (includes wire + heat shrink). Uses vendor mapping.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    data      = request.get_json()
    bid_data  = data.get("bid_data", {})
    calc_data = data.get("calc", {})
    quote_num = data.get("quote_num", f"AAE-{datetime.now().strftime('%Y%m%d')}")
    customer  = bid_data.get("customer_name", "")
    project   = bid_data.get("project_name", "")

    # Build structured BOM items from bid + calc
    bom_items = build_bom_items(bid_data, calc_data)
    # Convert to scan-BOM format — deterministic routing engine
    routing_rules = load_routing_rules()
    line_items = []
    for i, itm in enumerate(bom_items, 1):
        mfr = itm.get("manufacturer", "")
        pn  = itm.get("part_num", "")
        result = resolve_vendor(pn, mfr, routing_rules)
        line_items.append({
            "item_num":    i,
            "part_number": pn,
            "description": itm.get("description", ""),
            "qty":         itm.get("qty", 1),
            "unit":        itm.get("unit", "ea"),
            "manufacturer": result["manufacturer"],
            "vendor":      result["vendor"],
            "aae_cost":    itm.get("unit_cost", 0.0),
            "notes":       "",
            "category":    itm.get("category", "Other"),
        })

    # ── Excel workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master BOM"

    RED="9B1B1B"; DARK_RED="6B0A0A"; WHITE="FFFFFF"; LIGHT_RED="FDECEA"; MID_GRAY="F5F0F0"; DARK="2C2C2C"; TEAL="00897A"

    def s(cell, bold=False, bg=None, fg=WHITE, sz=10, ha="left"):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=False)

    thin = Side(style="thin", color="E0D0D0")
    bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)

    # 10 columns: A=Item,B=Part#,C=Description,D=Qty,E=Unit,F=Manufacturer,G=Vendor,H=Unit Cost,I=Total Cost,J=Notes
    for col, w in [("A",8),("B",22),("C",48),("D",6),("E",6),("F",22),("G",14),("H",14),("I",14),("J",20)]:
        ws.column_dimensions[col].width = w

    # Row 1: Banner
    ws.merge_cells("A1:J1"); ws.row_dimensions[1].height = 14
    ws["A1"] = "AAE AUTOMATION, INC.  |  UL-NNNY  |  UL-508A Certified Industrial Control Panel Specialists"
    s(ws["A1"], bold=True, bg=RED, sz=11, ha="center")

    # Row 2: Title | Quote#
    ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 34
    ws["A2"] = "BILL OF MATERIALS"
    ws["A2"].font      = Font(name="Arial", bold=True, color=WHITE, size=18)
    ws["A2"].fill      = PatternFill("solid", fgColor=DARK_RED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("F2:J2")
    ws["F2"] = quote_num
    s(ws["F2"], bold=True, bg=DARK_RED, fg="F4A9A8", sz=12, ha="right")

    # Row 3: Customer | Project
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:B3"); ws["A3"] = "Customer:"
    s(ws["A3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C3:E3"); ws["C3"] = customer
    s(ws["C3"], bg=LIGHT_RED, fg=DARK, sz=10)
    ws["F3"] = "Project:"
    s(ws["F3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G3:J3"); ws["G3"] = project
    s(ws["G3"], bg=LIGHT_RED, fg=DARK, sz=10)

    # Row 4: Date | Estimator
    ws.row_dimensions[4].height = 16
    ws.merge_cells("A4:B4"); ws["A4"] = "Date:"
    s(ws["A4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C4:E4"); ws["C4"] = datetime.now().strftime("%m/%d/%Y")
    s(ws["C4"], bg=MID_GRAY, fg=DARK, sz=9)
    ws["F4"] = "Estimator:"
    s(ws["F4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G4:J4"); ws["G4"] = bid_data.get("estimator_name", "AAE Automation")
    s(ws["G4"], bg=MID_GRAY, fg=DARK, sz=9)

    # Row 5: Internal notice
    ws.merge_cells("A5:J5"); ws.row_dimensions[5].height = 15
    ws["A5"] = "⚠  INTERNAL DOCUMENT ONLY — Not for Customer Distribution  ⚠"
    s(ws["A5"], bold=True, bg="FFF8E1", fg="CC6600", sz=9, ha="center")

    # Row 6: Column headers
    ws.row_dimensions[6].height = 20
    for ci, h in enumerate(["ITEM","PART NUMBER","DESCRIPTION","QTY","U/M","MANUFACTURER","VENDOR","UNIT COST","TOTAL COST","NOTES"], 1):
        c = ws.cell(row=6, column=ci, value=h)
        s(c, bold=True, bg=DARK, sz=9, ha="center")
        c.border = Border(bottom=Side(style="medium", color=RED))

    # Group by category
    grouped = defaultdict(list)
    cat_order = ["Enclosure","Power","Motor Ctrl","Control","PLC/Network","Terminals","Wiring","Other"]
    for item in line_items:
        cat = item.get("category", "Other")
        if cat not in cat_order: cat = "Other"
        grouped[cat].append(item)

    row = 7; item_counter = 0
    even_fill = PatternFill("solid", fgColor="FDF8F8")
    odd_fill  = PatternFill("solid", fgColor=WHITE)

    for cat in [c for c in cat_order if grouped[c]]:
        # Section header
        ws.merge_cells(f"A{row}:J{row}")
        hc = ws.cell(row=row, column=1, value=f"  {cat.upper()}")
        hc.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        hc.fill = PatternFill("solid", fgColor=RED)
        hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        for itm in grouped[cat]:
            item_counter += 1
            fill = even_fill if item_counter % 2 == 0 else odd_fill
            ws.row_dimensions[row].height = 15
            vals = [itm["item_num"], itm["part_number"], itm["description"],
                    itm["qty"], itm["unit"], itm["manufacturer"],
                    itm.get("vendor",""), itm["aae_cost"],
                    itm["qty"] * itm["aae_cost"],   # TOTAL COST = qty × unit cost
                    itm["notes"]]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = fill; c.border = bdr
                c.font = Font(name="Arial", size=9, color=DARK)
                if ci in (1, 4):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 7:  # Vendor — teal
                    c.font = Font(name="Arial", size=9, color=TEAL, bold=bool(val))
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 8:  # Unit Cost
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '"$"#,##0.000'
                    c.font = Font(name="Arial", size=9, color="555555")
                elif ci == 9:  # Total Cost — bold green
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '"$"#,##0.00'
                    c.font = Font(name="Arial", size=9, color="008800", bold=True)
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci==3))
            row += 1
        row += 1  # spacer

    # Total row
    ws.merge_cells(f"A{row}:G{row}")
    tl = ws.cell(row=row, column=1, value="TOTAL MATERIAL COST (AAE Cost):")
    tl.font = Font(name="Arial", bold=True, color=DARK_RED, size=10)
    tl.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tl.alignment = Alignment(horizontal="right", vertical="center")
    tv = ws.cell(row=row, column=9, value=f"=SUM(I7:I{row-1})")
    tv.font = Font(name="Arial", bold=True, color="008800", size=11)
    tv.number_format = '"$"#,##0.00'
    tv.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tv.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.row_dimensions[row].height = 20
    row += 2

    # Footer
    ws.merge_cells(f"A{row}:J{row}")
    ft = ws.cell(row=row, column=1,
        value="AAE Automation, Inc.  |  8528 SW 2nd St, Oklahoma City, OK 73128  |  405-210-1567  |  mfellers@aaeok.com")
    ft.font = Font(name="Arial", size=8, color="888080", italic=True)
    ft.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:J{row-3}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:6"

    # ── Build ZIP: Master BOM + one simple Excel file per vendor ───────────────
    import zipfile
    from collections import defaultdict as _vdd
    vendor_groups = _vdd(list)
    for _itm in line_items:
        _v = _itm.get("vendor", "") or "Unassigned"
        vendor_groups[_v].append(_itm)

    estimator  = bid_data.get("estimator_name", "")
    safe_proj  = (project or quote_num).replace(" ", "_").replace("/", "-")[:40]

    master_buf = io.BytesIO()
    wb.save(master_buf)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"AAE_Master_BOM_{safe_proj}.xlsx", master_buf.getvalue())
        for _vn in sorted(vendor_groups.keys()):
            if vendor_groups[_vn]:
                _vbuf = _build_vendor_excel(_vn, vendor_groups[_vn],
                                            customer, project, quote_num, estimator)
                _safe_vn = _vn.replace(" ", "_").replace("/", "-")[:20]
                zf.writestr(f"AAE_BOM_{_safe_vn}_{safe_proj}.xlsx", _vbuf.getvalue())
    zip_buf.seek(0)
    return send_file(zip_buf,
                     mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"AAE_BOM_Package_{safe_proj}.zip")

@app.route("/api/bid_quote_pdf/<int:bid_id>", methods=["GET"])
@require_auth
def bid_quote_pdf(bid_id):
    """Download quote PDF for a saved bid by ID."""
    try:
        if not _sb_url: return jsonify({"error": "Supabase not configured"}), 500
        sb = get_user_sb()
        result = sb.table("bids").select("*").eq("id", bid_id).single().execute()
        bid = result.data
        if not bid:
            return jsonify({"error": "Bid not found"}), 404
        bid_data  = json.loads(bid.get("bid_data", "{}"))
        calc_data = bid_data.get("calc", {})
        actual_bid = {k: v for k, v in bid_data.items() if k != "calc"}
        actual_bid["customer_name"] = bid.get("customer_name", "")
        actual_bid["project_name"]  = bid.get("project_name", "")
        actual_bid["estimator_name"]= bid.get("estimator_name", "")
        quote_num = f"AAE-{bid.get('created_at','')[:10].replace('-','')}-{bid_id}"
        pdf_buf   = generate_quote_pdf(actual_bid, calc_data, quote_num)
        return send_file(pdf_buf, mimetype="application/pdf", as_attachment=True,
                         download_name=f"AAE_Quote_{quote_num}.pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def build_bom_items(bid_data, calc_data):
    """Build a structured BOM item list from bid + calc data."""
    items = []
    def add(cat, part, desc, qty, unit, unit_cost, mfr=""):
        if qty <= 0: return
        items.append({"category": cat, "part_num": part, "description": desc,
                       "qty": qty, "unit": unit, "unit_cost": unit_cost,
                       "total_cost": qty * unit_cost, "manufacturer": mfr})

    enc  = int(bid_data.get("enc_qty", 1))
    add("Enclosure", "ENCLOSURE", "Panel Enclosure", enc, "ea", 450)
    add("Enclosure", "DIN-RAIL", "DIN Rail (1m stick)", int(bid_data.get("din_rail_runs", 0)), "ea", 12.50)
    add("Enclosure", "WIRE-DUCT", "Wire Duct Section", int(bid_data.get("wire_duct_runs", 0))*2, "ea", 8.75)
    add("Power", "MAIN-DISC", "Main Disconnect/Breaker", 1 if int(bid_data.get("main_amp",0))>0 else 0, "ea", 180)
    add("Power", "BREAKER-1P", "1P Branch Circuit Breaker", int(bid_data.get("branch_1p", 0)), "ea", 25)
    add("Power", "BREAKER-2P", "2P Branch Circuit Breaker", int(bid_data.get("branch_2p", 0)), "ea", 45)
    add("Power", "BREAKER-3P", "3P Branch Circuit Breaker", int(bid_data.get("branch_3p", 0)), "ea", 65)
    add("Power", "FUSED-DISC", "Fused Disconnect", int(bid_data.get("fused_disconnects", 0)), "ea", 85)
    add("Power", "CPT", "Control Power Transformer", 1 if str(bid_data.get("cpt_present","N")).upper()=="Y" else 0, "ea", 220)
    add("Power", "PWR-DIST-BLK", "Power Distribution Block", int(bid_data.get("pdb_qty", 0)), "ea", 28)
    add("Motor Ctrl", "CONT-SM", "Contactor ≤40A", int(bid_data.get("contactor_small", 0)), "ea", 65)
    add("Motor Ctrl", "CONT-LG", "Contactor 41-100A", int(bid_data.get("contactor_large", 0)), "ea", 140)
    add("Motor Ctrl", "OVERLOAD", "Overload Relay", int(bid_data.get("overload", 0)), "ea", 45)
    add("Motor Ctrl", "VFD-SM", "VFD ≤5HP", int(bid_data.get("vfd_small", 0)), "ea", 380)
    add("Motor Ctrl", "VFD-MD", "VFD 6-25HP", int(bid_data.get("vfd_med", 0)), "ea", 680)
    add("Motor Ctrl", "VFD-LG", "VFD 26-100HP", int(bid_data.get("vfd_large", 0)), "ea", 1400)
    add("Motor Ctrl", "SS-SM", "Soft Starter ≤50A", int(bid_data.get("soft_starter_small", 0)), "ea", 280)
    add("Motor Ctrl", "SS-LG", "Soft Starter >50A", int(bid_data.get("soft_starter_large", 0)), "ea", 520)
    add("Control", "RELAY-IC", "Ice Cube Relay", int(bid_data.get("relay_icecube", 0)), "ea", 18)
    add("Control", "RELAY-DIN", "DIN Mount Relay", int(bid_data.get("relay_din", 0)), "ea", 28)
    add("Control", "SSR", "Solid State Relay", int(bid_data.get("ssrs", 0)), "ea", 45)
    add("Control", "TIMER", "Timer Relay", int(bid_data.get("timers", 0)), "ea", 55)
    add("Control", "PILOT", "Pilot Light", int(bid_data.get("pilot_lights", 0)), "ea", 22)
    add("Control", "SEL-SW", "Selector Switch", int(bid_data.get("selectors", 0)), "ea", 28)
    add("Control", "PUSH-BTN", "Push Button", int(bid_data.get("push_buttons", 0)), "ea", 22)
    add("Control", "E-STOP", "E-Stop Button", int(bid_data.get("estops", 0)), "ea", 48)
    if str(bid_data.get("plc_present","N")).upper() == "Y":
        add("PLC/Network", "PLC-CTRL", "PLC Controller", 1, "ea", 600)
    add("PLC/Network", "DI-PT", "Digital Input Point", int(bid_data.get("plc_di", 0)), "pt", 8)
    add("PLC/Network", "DO-PT", "Digital Output Point", int(bid_data.get("plc_do", 0)), "pt", 10)
    add("PLC/Network", "AI-PT", "Analog Input Point", int(bid_data.get("plc_ai", 0)), "pt", 18)
    add("PLC/Network", "AO-PT", "Analog Output Point", int(bid_data.get("plc_ao", 0)), "pt", 22)
    if str(bid_data.get("hmi_present","N")).upper() == "Y":
        add("PLC/Network", "HMI", "HMI Display Panel", 1, "ea", 800)
    if str(bid_data.get("safety_relay","N")).upper() == "Y":
        add("PLC/Network", "SAFETY-CTRL", "Safety Relay/Controller", 1, "ea", 450)
    if str(bid_data.get("eth_switch","N")).upper() == "Y":
        add("PLC/Network", "ETH-SW", "Ethernet Switch", 1, "ea", 180)
    add("PLC/Network", "ETH-CBL", "Internal Ethernet Cable", int(bid_data.get("eth_cables", 0)), "ea", 12)
    add("Terminals", "TB-STD", "Standard Terminal Block", int(bid_data.get("tb_standard", 0)), "ea", 3.50, "Phoenix Contact")
    add("Terminals", "TB-GND", "Ground Terminal Block", int(bid_data.get("tb_ground", 0)), "ea", 4.20, "Phoenix Contact")
    add("Terminals", "TB-FUSED", "Fused Terminal Block", int(bid_data.get("tb_fused", 0)), "ea", 8.50, "Phoenix Contact")
    add("Terminals", "TB-DISC", "Disconnect Terminal Block", int(bid_data.get("tb_disconnect", 0)), "ea", 9.80)
    # Wire
    ctrl_ft = calc_data.get("ctrl_wire_ft", 0)
    pwr_ft  = calc_data.get("pwr_wire_ft", 0)
    wrate   = calc_data.get("wire_cost_per_ft", 0.40)
    if ctrl_ft + pwr_ft > 0:
        add("Wiring", "WIRE", f"Control/Power Wire ({ctrl_ft+pwr_ft} ft total)", ctrl_ft + pwr_ft, "ft", wrate, "Panduit")
    # Heat shrink labels
    hs_rolls = calc_data.get("hs_rolls", 0)
    if hs_rolls > 0:
        add("Wiring", "H075X044H1T", f"Heat Shrink Labels — {calc_data.get('hs_labels_qty',0)} pcs ({hs_rolls} roll{'s' if hs_rolls>1 else ''})", hs_rolls, "roll", 275, "Rexel")
    return items


def generate_bom_pdf(bid_data, calc_results, quote_number):
    """Generate a standalone BOM PDF with AAE branding."""
    buffer = io.BytesIO()
    from reportlab.lib.pagesizes import landscape
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    red   = colors.HexColor("#9B1B1B")
    dark  = colors.HexColor("#2C2C2C")
    lgray = colors.HexColor("#F0EDE8")
    styles = getSampleStyleSheet()

    story = []
    # Header
    hdr_data = [[Paragraph("<font color='white'><b>AAE AUTOMATION, INC.</b></font>",
                            ParagraphStyle("h", fontSize=16, fontName="Helvetica-Bold")),
                 Paragraph("<font color='white'>BILL OF MATERIALS</font>",
                            ParagraphStyle("h2", fontSize=12, fontName="Helvetica", alignment=TA_RIGHT))]]
    hdr_tbl = Table(hdr_data, colWidths=[5*inch, 4.5*inch])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), red),
        ("TOPPADDING",(0,0),(-1,-1), 10), ("BOTTOMPADDING",(0,0),(-1,-1), 10),
        ("LEFTPADDING",(0,0),(-1,-1), 14), ("RIGHTPADDING",(0,0),(-1,-1), 14),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 0.1*inch))

    # Sub-header info
    story.append(Paragraph(
        f"<b>Customer:</b> {bid_data.get('customer_name','')} &nbsp;&nbsp; "
        f"<b>Project:</b> {bid_data.get('project_name','')} &nbsp;&nbsp; "
        f"<b>Quote #:</b> {quote_number} &nbsp;&nbsp; "
        f"<b>Date:</b> {datetime.now().strftime('%m/%d/%Y')} &nbsp;&nbsp;&nbsp; "
        f"<font color='#CC6600'><b>⚠ INTERNAL DOCUMENT ONLY</b></font>",
        ParagraphStyle("sub", fontSize=9, fontName="Helvetica", spaceAfter=8)))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CCCCCC")))
    story.append(Spacer(1, 0.05*inch))

    # Column headers
    col_widths = [1.3*inch, 1.5*inch, 3.2*inch, 0.5*inch, 0.5*inch, 1*inch, 1*inch]
    col_heads  = [["Category","Part #","Description","Qty","Unit","Unit Cost","Total Cost"]]
    tbl_hdr = Table(col_heads, colWidths=col_widths)
    tbl_hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), dark),
        ("TEXTCOLOR",(0,0),(-1,-1), colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1), 8.5),
        ("TOPPADDING",(0,0),(-1,-1), 5), ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(tbl_hdr)

    # BOM rows
    bom_items = build_bom_items(bid_data, calc_results)
    last_cat  = None
    row_data  = []
    for item in bom_items:
        if item["category"] != last_cat:
            last_cat = item["category"]
            row_data.append([item["category"].upper(), "", "", "", "", "", ""])
        row_data.append([
            "", item["part_num"], item["description"],
            str(item["qty"]), item["unit"],
            f'${item["unit_cost"]:.2f}', f'${item["total_cost"]:.2f}'
        ])

    bom_tbl = Table(row_data, colWidths=col_widths)
    style_cmds = [
        ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1), 8.5),
        ("TOPPADDING",(0,0),(-1,-1), 4), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("LEFTPADDING",(0,0),(-1,-1), 6),
        ("LINEBELOW",(0,0),(-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("ALIGN",(3,0),(6,-1),"RIGHT"),
    ]
    # Category rows styling
    ri = 0
    for item in bom_items:
        if item["category"] != (bom_items[ri-1]["category"] if ri > 0 else None):
            style_cmds += [
                ("BACKGROUND",(0,ri),(6,ri), lgray),
                ("FONTNAME",(0,ri),(6,ri),"Helvetica-Bold"),
                ("TEXTCOLOR",(0,ri),(0,ri), red),
                ("SPAN",(0,ri),(6,ri)),
            ]
        elif ri % 2 == 0:
            style_cmds.append(("BACKGROUND",(0,ri),(6,ri), colors.HexColor("#FAFAFA")))
        ri += 1

    bom_tbl.setStyle(TableStyle(style_cmds))
    story.append(bom_tbl)

    # Total row
    total_raw = sum(i["total_cost"] for i in bom_items)
    story.append(Spacer(1, 0.1*inch))
    tot_data = [[Paragraph(f"<b>TOTAL MATERIAL COST (RAW): ${total_raw:,.2f}</b>",
                            ParagraphStyle("tot", fontSize=11, textColor=red, fontName="Helvetica-Bold", alignment=TA_RIGHT))]]
    story.append(Table(tot_data, colWidths=[9.5*inch]))
    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route("/api/bom_from_scan", methods=["POST"])
@require_auth
def bom_from_scan():
    """Generate a formatted AAE BOM Excel from scanned line items."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data       = request.get_json()
    line_items = data.get("bom_line_items", [])
    customer   = data.get("customer_name", "")
    project    = data.get("project_name", "")
    job_num    = data.get("job_number", f"AAE-{datetime.now().strftime('%Y%m%d')}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master BOM"

    # Column widths
    for col, w in [("A",10),("B",22),("C",54),("D",7),("E",7),("F",22),("G",14),("H",22)]:
        ws.column_dimensions[col].width = w

    RED      = "9B1B1B"; DARK_RED = "6B0A0A"; WHITE = "FFFFFF"
    LIGHT_RED= "FDECEA"; MID_GRAY = "F5F0F0"; DARK  = "2C2C2C"

    def s(cell, bold=False, bg=None, fg=WHITE, sz=10, ha="left", wrap=False):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)

    thin = Side(style="thin", color="E0D0D0")
    bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)

    # ── Column widths — 9 columns ──────────────────────────────────────────────
    for col, w in [("A",8),("B",22),("C",48),("D",6),("E",6),("F",22),("G",14),("H",12),("I",14),("J",20)]:
        ws.column_dimensions[col].width = w

    # ── Row 1: Banner (A1:I1) ──────────────────────────────────────────────────
    ws.merge_cells("A1:J1"); ws.row_dimensions[1].height = 14
    ws["A1"] = "AAE AUTOMATION, INC.  |  UL-NNNY  |  UL-508A Certified Industrial Control Panel Specialists"
    s(ws["A1"], bold=True, bg=RED, sz=11, ha="center")

    # ── Row 2: Title (A2:E2) | Job# (F2:I2) ───────────────────────────────────
    ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 34
    ws["A2"] = "BILL OF MATERIALS"
    ws["A2"].font      = Font(name="Arial", bold=True, color=WHITE, size=18)
    ws["A2"].fill      = PatternFill("solid", fgColor=DARK_RED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("F2:J2")
    ws["F2"] = job_num
    s(ws["F2"], bold=True, bg=DARK_RED, fg="F4A9A8", sz=12, ha="right")

    # ── Row 3: Customer (A3:B3 | C3:E3) | Project (F3 | G3:I3) ───────────────
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:B3"); ws["A3"] = "Customer:"
    s(ws["A3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C3:E3"); ws["C3"] = customer
    s(ws["C3"], bg=LIGHT_RED, fg=DARK, sz=10)
    ws["F3"] = "Project:"
    s(ws["F3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G3:J3"); ws["G3"] = project
    s(ws["G3"], bg=LIGHT_RED, fg=DARK, sz=10)

    # ── Row 4: Date (A4:B4 | C4:E4) | Estimator (F4 | G4:I4) ─────────────────
    ws.row_dimensions[4].height = 16
    ws.merge_cells("A4:B4"); ws["A4"] = "Date:"
    s(ws["A4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C4:E4"); ws["C4"] = datetime.now().strftime("%m/%d/%Y")
    s(ws["C4"], bg=MID_GRAY, fg=DARK, sz=9)
    ws["F4"] = "Estimator:"
    s(ws["F4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G4:J4"); ws["G4"] = "AAE Automation"
    s(ws["G4"], bg=MID_GRAY, fg=DARK, sz=9)

    # ── Row 5: Internal notice (A5:I5) ────────────────────────────────────────
    ws.merge_cells("A5:J5"); ws.row_dimensions[5].height = 15
    ws["A5"] = "⚠  INTERNAL DOCUMENT ONLY — Not for Customer Distribution  ⚠"
    s(ws["A5"], bold=True, bg="FFF8E1", fg="CC6600", sz=9, ha="center")

    # ── Row 6: Column headers ──────────────────────────────────────────────────
    ws.row_dimensions[6].height = 20
    for ci, h in enumerate(["ITEM","PART NUMBER","DESCRIPTION","QTY","U/M","MANUFACTURER","VENDOR","UNIT COST","TOTAL COST","NOTES"], 1):
        c = ws.cell(row=6, column=ci, value=h)
        s(c, bold=True, bg=DARK, sz=9, ha="center")
        c.border = Border(bottom=Side(style="medium", color=RED))

    # ── Load routing rules (deterministic 4-tier engine) ─────────────────────
    routing_rules = load_routing_rules()

    # ── Consolidate identical part numbers — sum quantities, preserve draw order ──
    from collections import OrderedDict as _OD
    _seen_pn = _OD()
    for _itm in sorted(line_items, key=lambda x: (x.get("item_num") or 9999)):
        _pn_key = (_itm.get("part_number") or "").strip().upper()
        if _pn_key and _pn_key != "[UNREADABLE]":
            if _pn_key in _seen_pn:
                # Duplicate part number — sum quantities, keep first row's other fields
                try:
                    _seen_pn[_pn_key]["qty"] = (
                        int(_seen_pn[_pn_key].get("qty", 1) or 1)
                        + int(_itm.get("qty", 1) or 1)
                    )
                except (ValueError, TypeError):
                    pass  # keep existing qty if int conversion fails
            else:
                _seen_pn[_pn_key] = dict(_itm)
        else:
            # Blank or unreadable part number — can't safely consolidate, keep as-is
            _unique_key = f"__nopn_{_itm.get('item_num', id(_itm))}"
            _seen_pn[_unique_key] = dict(_itm)
    line_items_out = list(_seen_pn.values())
    # ── End consolidation ──────────────────────────────────────────────────────

    row = 7; item_counter = 0
    even_fill = PatternFill("solid", fgColor="FDF8F8")
    odd_fill  = PatternFill("solid", fgColor=WHITE)
    thin      = Side(style="thin", color="E0D0D0")
    bdr       = Border(bottom=thin, left=thin, right=thin, top=thin)
    TEAL      = "00897A"

    # Write items in chronological BOM order (no category section headers)
    for itm in line_items_out:
        item_counter += 1
        fill = even_fill if item_counter % 2 == 0 else odd_fill
        ws.row_dimensions[row].height = 15
        mfr    = itm.get("manufacturer", "")
        pn     = itm.get("part_number", "")
        result = resolve_vendor(pn, mfr, routing_rules)
        vendor = result["vendor"]
        itm["_resolved_vendor"] = vendor  # store for vendor sheet grouping
        unit_cost = float(itm.get("aae_cost", 0) or 0)
        qty_raw = itm.get("qty", 1) or 1
        try:
            qty_val = int(qty_raw)
        except (ValueError, TypeError):
            qty_val = 1
        vals = [
            item_counter,                          # A: sequential ITEM #
            itm.get("part_number", ""),            # B: PART NUMBER
            itm.get("description", ""),            # C: DESCRIPTION
            qty_val,                               # D: QTY (consolidated)
            itm.get("unit", "ea"),                 # E: U/M
            mfr,                                   # F: MANUFACTURER
            vendor,                                # G: VENDOR
            unit_cost,                             # H: UNIT COST
            unit_cost * qty_val,                   # I: TOTAL COST
            itm.get("notes", "")                   # J: NOTES
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill; c.border = bdr
            c.font = Font(name="Arial", size=9, color=DARK)
            if ci in (1, 4):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 7:  # Vendor — teal, centered
                c.font = Font(name="Arial", size=9, color=TEAL, bold=bool(val))
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (8, 9):  # Unit Cost / Total Cost
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '"$"#,##0.00'
                c.font = Font(name="Arial", size=9, color="AAAAAA", italic=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci==3))
        row += 1

    # Total row (spans A:G, value in H)
    ws.merge_cells(f"A{row}:H{row}")
    tl = ws.cell(row=row, column=1, value="TOTAL MATERIAL COST (pricing TBD from QuickBooks):")
    tl.font = Font(name="Arial", bold=True, color=DARK_RED, size=10)
    tl.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tl.alignment = Alignment(horizontal="right", vertical="center")
    tv = ws.cell(row=row, column=9, value=f"=SUM(I7:I{row-1})")
    tv.font = Font(name="Arial", bold=True, color="AAAAAA", size=11, italic=True)
    tv.number_format = '"$"#,##0.00'
    tv.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tv.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.row_dimensions[row].height = 20
    row += 1

    # QB note
    ws.merge_cells(f"A{row}:J{row}")
    note = ws.cell(row=row, column=1,
        value="NOTE: AAE Cost column will be populated from QuickBooks pricing. Vendor column auto-assigned from AAE vendor database.")
    note.font = Font(name="Arial", size=8, color="888080", italic=True)
    note.alignment = Alignment(horizontal="left")
    row += 2

    # Footer
    ws.merge_cells(f"A{row}:J{row}")
    ft = ws.cell(row=row, column=1,
        value="AAE Automation, Inc.  |  8528 SW 2nd St, Oklahoma City, OK 73128  |  405-210-1567  |  mfellers@aaeok.com")
    ft.font = Font(name="Arial", size=8, color="888080", italic=True)
    ft.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:J{row-3}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:6"

    # ── Build ZIP: Master BOM + one simple Excel file per vendor ───────────────
    import zipfile
    from collections import defaultdict as _vdd2
    scan_vendor_groups = _vdd2(list)
    for _itm in line_items_out:
        _v = _itm.get("_resolved_vendor", "") or "Unassigned"
        _norm = dict(_itm); _norm["vendor"] = _v
        scan_vendor_groups[_v].append(_norm)

    safe_proj = (project or job_num).replace(" ", "_").replace("/", "-")[:40]

    master_buf = io.BytesIO()
    wb.save(master_buf)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"AAE_Master_BOM_{safe_proj}.xlsx", master_buf.getvalue())
        for _vn in sorted(scan_vendor_groups.keys()):
            if scan_vendor_groups[_vn]:
                _vbuf = _build_vendor_excel(_vn, scan_vendor_groups[_vn],
                                            customer, project, job_num, "")
                _safe_vn = _vn.replace(" ", "_").replace("/", "-")[:20]
                zf.writestr(f"AAE_BOM_{_safe_vn}_{safe_proj}.xlsx", _vbuf.getvalue())
    zip_buf.seek(0)
    return send_file(zip_buf,
                     mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"AAE_BOM_Package_{safe_proj}.zip")



# ── Vendor BOM File Builder ────────────────────────────────────────────────
def _build_vendor_excel(vendor_name, items, customer, project, quote_num, estimator=""):
    """Create a standalone simple Excel workbook for one vendor's BOM.
    6 columns: ITEM, PART NUMBER, DESCRIPTION, QTY, MANUFACTURER, VENDOR
    No cost columns — clean purchase-order format for sending to vendors.
    Returns an io.BytesIO buffer ready for writing into a ZIP.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vendor_name[:31].replace("/","&").replace("\\","&").replace("?","").replace("*","")

    RED = "9B1B1B"; DARK_RED = "6B0A0A"; WHITE = "FFFFFF"
    LIGHT_RED = "FDECEA"; MID_GRAY = "F5F0F0"; DARK = "2C2C2C"; TEAL = "00897A"

    def s(cell, bold=False, bg=None, fg=WHITE, sz=10, ha="left"):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=False)

    thin = Side(style="thin", color="E0D0D0")
    bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)

    # 8 columns: Item, Part Number, Description, Qty, U/M, Manufacturer, Vendor, Notes
    for col, w in [("A",8),("B",24),("C",52),("D",8),("E",8),("F",26),("G",18),("H",24)]:
        ws.column_dimensions[col].width = w

    # Row 1: Banner
    ws.merge_cells("A1:H1"); ws.row_dimensions[1].height = 14
    ws["A1"] = f"AAE AUTOMATION — VENDOR ORDER: {vendor_name.upper()}  |  UL-508A Certified Industrial Control Panel Specialists"
    s(ws["A1"], bold=True, bg=RED, sz=11, ha="center")

    # Row 2: Title | Quote#
    ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 34
    ws["A2"] = f"PURCHASE ORDER — {vendor_name.upper()}"
    ws["A2"].font = Font(name="Arial", bold=True, color=WHITE, size=16)
    ws["A2"].fill = PatternFill("solid", fgColor=DARK_RED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("F2:H2")
    ws["F2"] = quote_num
    s(ws["F2"], bold=True, bg=DARK_RED, fg="F4A9A8", sz=12, ha="right")

    # Row 3: Customer | Project
    ws.row_dimensions[3].height = 18
    ws["A3"] = "Customer:"
    s(ws["A3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("B3:C3"); ws["B3"] = customer
    s(ws["B3"], bg=LIGHT_RED, fg=DARK, sz=10)
    ws["D3"] = "Project:"
    s(ws["D3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("E3:H3"); ws["E3"] = project
    s(ws["E3"], bg=LIGHT_RED, fg=DARK, sz=10)

    # Row 4: Date | Vendor
    ws.row_dimensions[4].height = 16
    ws["A4"] = "Date:"
    s(ws["A4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("B4:C4"); ws["B4"] = datetime.now().strftime("%m/%d/%Y")
    s(ws["B4"], bg=MID_GRAY, fg=DARK, sz=9)
    ws["D4"] = "Send To:"
    s(ws["D4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("E4:H4"); ws["E4"] = vendor_name
    s(ws["E4"], bg=MID_GRAY, fg=DARK, sz=9)

    # Row 5: Column headers
    ws.row_dimensions[5].height = 20
    for ci, h in enumerate(["ITEM", "PART NUMBER", "DESCRIPTION", "QTY", "U/M", "MANUFACTURER", "VENDOR", "NOTES"], 1):
        c = ws.cell(row=5, column=ci, value=h)
        s(c, bold=True, bg=DARK, sz=9, ha="center")
        c.border = Border(bottom=Side(style="medium", color=RED))

    # Data rows — flat list, no category groupings
    row = 6
    even_fill = PatternFill("solid", fgColor="FDF8F8")
    odd_fill  = PatternFill("solid", fgColor=WHITE)

    for i, itm in enumerate(items, 1):
        fill = even_fill if i % 2 == 0 else odd_fill
        ws.row_dimensions[row].height = 15
        vals = [
            i,
            itm.get("part_number") or itm.get("part_num", ""),
            itm.get("description", ""),
            itm.get("qty", 1),
            itm.get("unit", "ea"),
            itm.get("manufacturer", ""),
            itm.get("vendor", "") or vendor_name,
            itm.get("notes", ""),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill; c.border = bdr
            c.font = Font(name="Arial", size=9, color=DARK)
            if ci in (1, 4):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:  # U/M
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 7:  # Vendor — teal accent
                c.font = Font(name="Arial", size=9, color=TEAL, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci == 3))
        row += 1

    # Total row
    total_qty = sum(int(itm.get("qty") or 1) for itm in items)
    ws.merge_cells(f"A{row}:C{row}")
    tl = ws.cell(row=row, column=1, value=f"TOTAL LINE ITEMS — {vendor_name.upper()}:")
    tl.font = Font(name="Arial", bold=True, color=DARK_RED, size=10)
    tl.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tl.alignment = Alignment(horizontal="right", vertical="center")
    tv = ws.cell(row=row, column=4, value=total_qty)
    tv.font = Font(name="Arial", bold=True, color="008800", size=11)
    tv.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tv.alignment = Alignment(horizontal="center", vertical="center")
    for ci in range(5, 9):
        ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.row_dimensions[row].height = 20
    row += 2

    # Footer
    ws.merge_cells(f"A{row}:H{row}")
    ft = ws.cell(row=row, column=1,
        value="AAE Automation, Inc.  |  8528 SW 2nd St, Oklahoma City, OK 73128  |  405-210-1567  |  mfellers@aaeok.com")
    ft.font = Font(name="Arial", size=8, color="888080", italic=True)
    ft.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{row - 3}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:5"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Vendor BOM Sheet Helper (legacy — kept for reference) ─────────────────
def _write_vendor_bom_sheet(wb, vendor_name, items, customer, project, quote_num,
                             estimator="", show_aae_cost=True):
    """
    Add a vendor-specific BOM worksheet to an existing openpyxl Workbook.
    Mirrors the master BOM format, filtered to one vendor's items.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    # Sheet name: Excel limits to 31 chars, strip invalid chars
    safe_name = vendor_name[:31].replace("/","&").replace("\\","&").replace("?","").replace("*","")
    ws = wb.create_sheet(title=safe_name)

    RED="9B1B1B"; DARK_RED="6B0A0A"; WHITE="FFFFFF"; LIGHT_RED="FDECEA"
    MID_GRAY="F5F0F0"; DARK="2C2C2C"; TEAL="00897A"

    def s(cell, bold=False, bg=None, fg=WHITE, sz=10, ha="left"):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=False)

    thin = Side(style="thin", color="E0D0D0")
    bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)

    for col, w in [("A",8),("B",22),("C",48),("D",6),("E",6),("F",22),("G",14),("H",14),("I",14),("J",20)]:
        ws.column_dimensions[col].width = w

    # Row 1: Banner
    ws.merge_cells("A1:J1"); ws.row_dimensions[1].height = 14
    ws["A1"] = f"AAE AUTOMATION — VENDOR ORDER: {vendor_name.upper()}  |  UL-508A Certified Panel Specialists"
    s(ws["A1"], bold=True, bg=RED, sz=11, ha="center")

    # Row 2: Title | Quote#
    ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 34
    ws["A2"] = f"PURCHASE ORDER — {vendor_name.upper()}"
    ws["A2"].font      = Font(name="Arial", bold=True, color=WHITE, size=16)
    ws["A2"].fill      = PatternFill("solid", fgColor=DARK_RED)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("F2:J2")
    ws["F2"] = quote_num
    s(ws["F2"], bold=True, bg=DARK_RED, fg="F4A9A8", sz=12, ha="right")

    # Row 3: Customer | Project
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:B3"); ws["A3"] = "Customer:"
    s(ws["A3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C3:E3"); ws["C3"] = customer
    s(ws["C3"], bg=LIGHT_RED, fg=DARK, sz=10)
    ws["F3"] = "Project:"
    s(ws["F3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G3:J3"); ws["G3"] = project
    s(ws["G3"], bg=LIGHT_RED, fg=DARK, sz=10)

    # Row 4: Date | Estimator
    ws.row_dimensions[4].height = 16
    ws.merge_cells("A4:B4"); ws["A4"] = "Date:"
    s(ws["A4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("C4:E4"); ws["C4"] = datetime.now().strftime("%m/%d/%Y")
    s(ws["C4"], bg=MID_GRAY, fg=DARK, sz=9)
    ws["F4"] = "Send To:"
    s(ws["F4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
    ws.merge_cells("G4:J4"); ws["G4"] = vendor_name
    s(ws["G4"], bg=MID_GRAY, fg=DARK, sz=9)

    # Row 5: Vendor notice
    ws.merge_cells("A5:J5"); ws.row_dimensions[5].height = 15
    ws["A5"] = f"⚠  VENDOR PURCHASE ORDER — Send to: {vendor_name}  |  Estimator: {estimator or 'AAE Automation'}  ⚠"
    s(ws["A5"], bold=True, bg="FFF8E1", fg="CC6600", sz=9, ha="center")

    # Row 6: Column headers
    ws.row_dimensions[6].height = 20
    for ci, h in enumerate(["ITEM","PART NUMBER","DESCRIPTION","QTY","U/M","MANUFACTURER","VENDOR","UNIT COST","TOTAL COST","NOTES"], 1):
        c = ws.cell(row=6, column=ci, value=h)
        s(c, bold=True, bg=DARK, sz=9, ha="center")
        c.border = Border(bottom=Side(style="medium", color=RED))

    # Group items by category
    grouped = defaultdict(list)
    cat_order = ["Enclosure","Power","Motor Ctrl","Control","Control Devices","PLC/Network",
                 "Terminals","Relays","HMI/Computer","Wiring","Markers","Other"]
    for item in items:
        cat = item.get("category","Other")
        if cat not in cat_order:
            cat_order.append(cat)
        grouped[cat].append(item)

    row = 7; item_counter = 0
    even_fill = PatternFill("solid", fgColor="FDF8F8")
    odd_fill  = PatternFill("solid", fgColor=WHITE)

    for cat in [c for c in cat_order if grouped[c]]:
        ws.merge_cells(f"A{row}:J{row}")
        hc = ws.cell(row=row, column=1, value=f"  {cat.upper()}")
        hc.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        hc.fill = PatternFill("solid", fgColor=RED)
        hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        for itm in grouped[cat]:
            item_counter += 1
            fill = even_fill if item_counter % 2 == 0 else odd_fill
            ws.row_dimensions[row].height = 15
            unit_cost = float(itm.get("aae_cost") or itm.get("unit_cost") or 0)
            qty = itm.get("qty", 1)
            vals = [
                itm.get("item_num", item_counter),
                itm.get("part_number") or itm.get("part_num",""),
                itm.get("description",""),
                qty,
                itm.get("unit","ea"),
                itm.get("manufacturer",""),
                itm.get("vendor","") or vendor_name,
                unit_cost if show_aae_cost else 0.00,
                qty * unit_cost if show_aae_cost else 0,
                itm.get("notes",""),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = fill; c.border = bdr
                c.font = Font(name="Arial", size=9, color=DARK)
                if ci in (1, 4):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 7:
                    c.font = Font(name="Arial", size=9, color=TEAL, bold=True)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif ci == 8:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '"$"#,##0.000'
                    c.font = Font(name="Arial", size=9,
                                  color="555555" if show_aae_cost else "AAAAAA",
                                  italic=not show_aae_cost)
                elif ci == 9:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '"$"#,##0.00'
                    c.font = Font(name="Arial", size=9,
                                  color="008800" if show_aae_cost else "AAAAAA",
                                  bold=show_aae_cost, italic=not show_aae_cost)
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci==3))
            row += 1
        row += 1  # spacer between categories

    # Total row
    ws.merge_cells(f"A{row}:G{row}")
    tl = ws.cell(row=row, column=1, value=f"TOTAL — {vendor_name.upper()} ORDER:")
    tl.font = Font(name="Arial", bold=True, color=DARK_RED, size=10)
    tl.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tl.alignment = Alignment(horizontal="right", vertical="center")
    tv = ws.cell(row=row, column=9, value=f"=SUM(I7:I{row-1})")
    tv.font = Font(name="Arial", bold=True,
                   color="008800" if show_aae_cost else "AAAAAA",
                   size=11, italic=not show_aae_cost)
    tv.number_format = '"$"#,##0.00'
    tv.fill = PatternFill("solid", fgColor=LIGHT_RED)
    tv.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=8).fill  = PatternFill("solid", fgColor=LIGHT_RED)
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=LIGHT_RED)
    ws.row_dimensions[row].height = 20
    row += 2

    # Footer
    ws.merge_cells(f"A{row}:J{row}")
    ft = ws.cell(row=row, column=1,
        value="AAE Automation, Inc.  |  8528 SW 2nd St, Oklahoma City, OK 73128  |  405-210-1567  |  mfellers@aaeok.com")
    ft.font = Font(name="Arial", size=8, color="888080", italic=True)
    ft.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:J{row-3}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:6"
    return ws


# ═══════════════════════════════════════════════════════════════════
# ADMIN API ROUTES
# ═══════════════════════════════════════════════════════════════════
# ADMIN API ROUTES — All require server-verified JWT auth
# ═══════════════════════════════════════════════════════════════════

# ── Labor Rates (admin write, authenticated read) ─────────────────────────────
@app.route("/api/labor_rates", methods=["GET"])
@require_auth
def get_labor_rates_api():
    if not supabase:
        cats = {
            "enclosure_prep":"Enclosure","subpanel_mount":"Enclosure","panel_layout":"Enclosure",
            "din_rail":"Enclosure","wire_duct":"Enclosure","enc_accessory":"Enclosure","door_component":"Enclosure",
            "main_breaker_small":"Power","main_breaker_large":"Power","branch_breaker_1p":"Power",
            "branch_breaker_23p":"Power","fused_disconnect":"Power","cpt":"Power","pdb":"Power",
            "relay_icecube":"Motor Ctrl","relay_din":"Motor Ctrl","contactor_small":"Motor Ctrl",
            "contactor_large":"Motor Ctrl","overload":"Motor Ctrl","timer":"Motor Ctrl","ssr":"Motor Ctrl",
            "vfd_small":"Motor Ctrl","vfd_med":"Motor Ctrl","vfd_large":"Motor Ctrl",
            "soft_starter_small":"Motor Ctrl","soft_starter_large":"Motor Ctrl",
            "pilot_light":"Control","selector":"Control","pushbutton":"Control","estop":"Control",
            "plc_rack":"PLC/Network","plc_di_do":"PLC/Network","plc_ai_ao":"PLC/Network",
            "hmi":"PLC/Network","safety_relay":"PLC/Network","eth_switch":"PLC/Network","eth_cable":"PLC/Network",
            "tb_standard":"Terminals","tb_ground":"Terminals","tb_fused":"Terminals",
            "tb_disconnect":"Terminals","tb_accessories":"Terminals","terminal_markers":"Terminals",
            "wire_land_control":"Wiring","ferrule":"Wiring","wire_route":"Wiring",
            "heat_shrink_label":"Wiring","heat_shrink_batch":"Wiring",
            "ul_labels":"UL/QC","continuity_check":"UL/QC","hipot":"UL/QC","as_built":"UL/QC","qc_signoff":"UL/QC",
        }
        return jsonify([{"rate_key":k,"rate_value":v,"category":cats.get(k,"Other"),"description":k.replace("_"," ").title()}
                        for k,v in LABOR_RATES.items()])
    try:
        sb = get_user_sb()
        rows = sb.table("aae_labor_rates").select("*").order("category").order("rate_key").execute()
        return jsonify(rows.data)
    except Exception as e:
        print(f"labor_rates DB error (returning defaults): {e}")
        cats = {
            "enclosure_prep":"Enclosure","subpanel_mount":"Enclosure","panel_layout":"Enclosure",
            "din_rail":"Enclosure","wire_duct":"Enclosure","enc_accessory":"Enclosure","door_component":"Enclosure",
            "main_breaker_small":"Power","main_breaker_large":"Power","branch_breaker_1p":"Power",
            "branch_breaker_23p":"Power","fused_disconnect":"Power","cpt":"Power","pdb":"Power",
            "relay_icecube":"Motor Ctrl","relay_din":"Motor Ctrl","contactor_small":"Motor Ctrl",
            "contactor_large":"Motor Ctrl","overload":"Motor Ctrl","timer":"Motor Ctrl","ssr":"Motor Ctrl",
            "vfd_small":"Motor Ctrl","vfd_med":"Motor Ctrl","vfd_large":"Motor Ctrl",
            "soft_starter_small":"Motor Ctrl","soft_starter_large":"Motor Ctrl",
            "pilot_light":"Control","selector":"Control","pushbutton":"Control","estop":"Control",
            "plc_rack":"PLC/Network","plc_di_do":"PLC/Network","plc_ai_ao":"PLC/Network",
            "hmi":"PLC/Network","safety_relay":"PLC/Network","eth_switch":"PLC/Network","eth_cable":"PLC/Network",
            "tb_standard":"Terminals","tb_ground":"Terminals","tb_fused":"Terminals",
            "tb_disconnect":"Terminals","tb_accessories":"Terminals","terminal_markers":"Terminals",
            "wire_land_control":"Wiring","ferrule":"Wiring","wire_route":"Wiring",
            "heat_shrink_label":"Wiring","heat_shrink_batch":"Wiring",
            "ul_labels":"UL/QC","continuity_check":"UL/QC","hipot":"UL/QC","as_built":"UL/QC","qc_signoff":"UL/QC",
        }
        return jsonify([{"rate_key":k,"rate_value":v,"category":cats.get(k,"Other"),"description":k.replace("_"," ").title()}
                        for k,v in LABOR_RATES.items()])

@app.route("/api/labor_rates/<rate_key>", methods=["PUT"])
@require_role("admin", "accounting", require_mfa=True)
def update_labor_rate(rate_key):
    data = request.get_json()
    if not supabase: return jsonify({"error":"DB not configured"}), 500
    try:
        new_val = float(data["rate_value"])
        updates = {"rate_value": new_val, "updated_by": g.user["email"], "updated_at": datetime.now().isoformat()}
        sb = get_user_sb()
        sb.table("aae_labor_rates").update(updates).eq("rate_key", rate_key).execute()
        audit_log("update_labor_rate", "aae_labor_rates", rate_key, {"rate_value": new_val})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/labor_rates/bulk", methods=["POST"])
@require_role("admin", "accounting", require_mfa=True)
def bulk_update_labor_rates():
    """Update multiple rates at once — admin + MFA required."""
    data = request.get_json()
    updates = data.get("rates", {})
    if not supabase:
        return jsonify({"success": True, "updated": 0, "note": "DB not configured"})
    try:
        now = datetime.now().isoformat()
        for key, val in updates.items():
            sb = get_user_sb()
            sb.table("aae_labor_rates").update(
                {"rate_value": float(val), "updated_by": g.user["email"], "updated_at": now}
            ).eq("rate_key", key).execute()
        audit_log("bulk_update_labor_rates", "aae_labor_rates", None, {"count": len(updates), "keys": list(updates.keys())})
        return jsonify({"success": True, "updated": len(updates)})
    except Exception as e:
        print(f"bulk_update_labor_rates DB error: {e}")
        return jsonify({"error": str(e)}), 500

# ── Vendors ───────────────────────────────────────────────────────────
@app.route("/api/vendors", methods=["GET"])
@require_auth
def get_vendors():
    if not supabase:
        return jsonify([])
    try:
        sb = get_user_sb()
        rows = sb.table("aae_vendors").select("*").order("vendor_name").order("manufacturer").execute()
        return jsonify(rows.data)
    except Exception as e:
        print(f"vendors DB error (returning empty): {e}")
        return jsonify([])

@app.route("/api/vendors", methods=["POST"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def create_vendor():
    data = request.get_json()
    if not supabase:
        return jsonify({"error": "DB not configured"}), 503
    try:
        sb = get_user_sb()
        result = sb.table("aae_vendors").insert({
            "vendor_name": data["vendor_name"], "manufacturer": data["manufacturer"],
            "account_number": data.get("account_number", ""),
            "notes": data.get("notes", ""), "updated_by": g.user["email"]
        }).execute()
        new_id = result.data[0]["id"]
        audit_log("create_vendor", "aae_vendors", new_id, {"manufacturer": data["manufacturer"], "vendor": data["vendor_name"]})
        return jsonify({"success": True, "vendor": result.data[0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vendors/<int:vid>", methods=["PUT"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def update_vendor(vid):
    data = request.get_json()
    if not supabase: return jsonify({"error": "DB not configured"}), 500
    try:
        updates = {k: data[k] for k in ["vendor_name","manufacturer","account_number","notes","active"] if k in data}
        updates["updated_at"] = datetime.now().isoformat()
        updates["updated_by"] = g.user["email"]
        sb = get_user_sb()
        sb.table("aae_vendors").update(updates).eq("id", vid).execute()
        audit_log("update_vendor", "aae_vendors", vid, updates)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vendors/<int:vid>", methods=["DELETE"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def delete_vendor(vid):
    if not supabase: return jsonify({"error": "DB not configured"}), 500
    try:
        sb = get_user_sb()
        sb.table("aae_vendors").update({"active": False, "updated_at": datetime.now().isoformat()}).eq("id", vid).execute()
        audit_log("delete_vendor", "aae_vendors", vid)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Vendor Routing: Part Override Routes ─────────────────────────────────────
@app.route("/api/part_overrides", methods=["GET"])
@require_auth
def get_part_overrides():
    """List all active part-level vendor overrides."""
    if not supabase:
        return jsonify([])
    try:
        sb = get_user_sb()
        rows = sb.table("aae_vendor_part_overrides") \
            .select("*").eq("active", True) \
            .order("part_number").execute()
        return jsonify(rows.data)
    except Exception as e:
        print(f"part_overrides DB error: {e}")
        return jsonify([])

@app.route("/api/part_overrides", methods=["POST"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def create_part_override():
    """Create a new part-level vendor override (Tier 1)."""
    data = request.get_json()
    if not supabase:
        return jsonify({"error": "DB not configured"}), 503
    part_num = (data.get("part_number") or "").strip()
    vendor   = (data.get("vendor_name") or "").strip()
    if not part_num or not vendor:
        return jsonify({"error": "Part number and vendor required"}), 400
    try:
        sb = get_user_sb()
        result = sb.table("aae_vendor_part_overrides").insert({
            "part_number":  part_num,
            "manufacturer": data.get("manufacturer", ""),
            "vendor_name":  vendor,
            "notes":        data.get("notes", ""),
            "updated_by":   g.user["email"],
        }).execute()
        new_id = result.data[0]["id"]
        audit_log("create_part_override", "aae_vendor_part_overrides", new_id,
                  {"part_number": part_num, "vendor": vendor})
        return jsonify({"success": True, "override": result.data[0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/part_overrides/<int:oid>", methods=["DELETE"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def delete_part_override(oid):
    """Soft-delete a part override."""
    if not supabase:
        return jsonify({"error": "DB not configured"}), 500
    try:
        sb = get_user_sb()
        sb.table("aae_vendor_part_overrides") \
            .update({"active": False, "updated_at": datetime.now().isoformat()}) \
            .eq("id", oid).execute()
        audit_log("delete_part_override", "aae_vendor_part_overrides", oid)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Vendor Routing: Prefix / Family Rule Routes ─────────────────────────────
@app.route("/api/prefix_rules", methods=["GET"])
@require_auth
def get_prefix_rules():
    """List all active prefix/family vendor rules."""
    if not supabase:
        return jsonify([])
    try:
        sb = get_user_sb()
        rows = sb.table("aae_vendor_prefix_rules") \
            .select("*").eq("active", True) \
            .order("priority", desc=True).order("prefix").execute()
        return jsonify(rows.data)
    except Exception as e:
        print(f"prefix_rules DB error: {e}")
        return jsonify([])

@app.route("/api/prefix_rules", methods=["POST"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def create_prefix_rule():
    """Create a new prefix/family vendor rule (Tier 2).
    Auto-generates the regex pattern from the user-supplied prefix."""
    data = request.get_json()
    if not supabase:
        return jsonify({"error": "DB not configured"}), 503
    prefix = (data.get("prefix") or "").strip()
    vendor = (data.get("vendor_name") or "").strip()
    if not prefix or not vendor:
        return jsonify({"error": "Prefix and vendor required"}), 400
    # Auto-generate regex: escape special chars, anchor to start
    escaped = re.escape(prefix)
    regex_pattern = f"^{escaped}"
    # Longer prefixes get higher priority (more specific match wins)
    priority = data.get("priority", len(prefix) * 10)
    try:
        sb = get_user_sb()
        result = sb.table("aae_vendor_prefix_rules").insert({
            "prefix":        prefix,
            "regex_pattern": regex_pattern,
            "manufacturer":  data.get("manufacturer", ""),
            "vendor_name":   vendor,
            "notes":         data.get("notes", ""),
            "priority":      priority,
            "updated_by":    g.user["email"],
        }).execute()
        new_id = result.data[0]["id"]
        audit_log("create_prefix_rule", "aae_vendor_prefix_rules", new_id,
                  {"prefix": prefix, "vendor": vendor})
        return jsonify({"success": True, "rule": result.data[0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/prefix_rules/<int:rid>", methods=["DELETE"])
@require_role("admin", "purchasing", "accounting", require_mfa=True)
def delete_prefix_rule(rid):
    """Soft-delete a prefix rule."""
    if not supabase:
        return jsonify({"error": "DB not configured"}), 500
    try:
        sb = get_user_sb()
        sb.table("aae_vendor_prefix_rules") \
            .update({"active": False, "updated_at": datetime.now().isoformat()}) \
            .eq("id", rid).execute()
        audit_log("delete_prefix_rule", "aae_vendor_prefix_rules", rid)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Vendor Routing: Test Endpoint ────────────────────────────────────────────
@app.route("/api/test_routing", methods=["POST"])
@require_auth
def test_routing():
    """Test where a part number would route through the 4-tier engine.
    Returns: {vendor, manufacturer, note, matched_tier}"""
    data = request.get_json()
    part_number  = data.get("part_number", "")
    manufacturer = data.get("manufacturer", "")
    rules  = load_routing_rules()
    result = resolve_vendor(part_number, manufacturer, rules)
    return jsonify(result)


# ── Hour Breakdown Report (Excel download) ────────────────────────────
@app.route("/api/hours_report", methods=["POST"])
@require_auth
def hours_report():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    data      = request.get_json()
    calc      = data.get("calc", {})
    bid       = data.get("bid_data", {})
    breakdown = calc.get("hour_breakdown", {})
    project   = bid.get("project_name","Unknown Project")
    customer  = bid.get("customer_name","")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Labor Hour Breakdown"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40

    RED="9B1B1B"; DARK_RED="6B0A0A"; WHITE="FFFFFF"; LIGHT_RED="FDECEA"
    DARK="2C2C2C"; GRAY="F5F0F0"

    def hdr(row, vals, bg, fg=WHITE, bold=True, sz=10):
        for ci, v in enumerate(vals, 1):

            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
            if bg: c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center", indent=(1 if ci==1 else 0))

    def row_data(r, label, hrs, rate_key=None, qty=None, rate=None, note=""):
        thin = Side(style="thin", color="E0D0D0")
        bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)
        vals = [label, f"{hrs:.2f} hrs" if isinstance(hrs,(int,float)) else hrs,
                f"{rate:.2f} min/ea × {qty}" if (rate and qty is not None) else "", note]
        even = r%2==0
        fill = PatternFill("solid", fgColor=GRAY if even else WHITE)
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill; c.border = bdr
            c.font = Font(name="Arial", size=9, color=DARK)
            c.alignment = Alignment(horizontal="right" if ci==2 else "left", vertical="center")

    # Banner
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height = 14
    c=ws["A1"]; c.value="AAE AUTOMATION — LABOR HOUR BREAKDOWN REPORT"; c.fill=PatternFill("solid",fgColor=RED)
    c.font=Font(name="Arial",bold=True,color=WHITE,size=11); c.alignment=Alignment(horizontal="center",vertical="center")

    # Project header
    ws.merge_cells("A2:B2"); ws["A2"]=f"Project: {project}"
    ws["A2"].font=Font(name="Arial",bold=True,color=DARK_RED,size=10); ws["A2"].fill=PatternFill("solid",fgColor=LIGHT_RED)
    ws.merge_cells("C2:D2"); ws["C2"]=f"Customer: {customer}"
    ws["C2"].font=Font(name="Arial",size=9,color=DARK); ws["C2"].fill=PatternFill("solid",fgColor=LIGHT_RED)
    ws.row_dimensions[2].height=18

    ws.merge_cells("A3:B3"); ws["A3"]=f"Date: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    ws["A3"].font=Font(name="Arial",size=9,color=DARK); ws["A3"].fill=PatternFill("solid",fgColor=GRAY)
    ws.merge_cells("C3:D3")
    ws["C3"]=f"Complexity: {bid.get('complexity','STANDARD')}  |  Tech: {bid.get('tech_level','JOURNEYMAN')}"
    ws["C3"].font=Font(name="Arial",size=9,color=DARK); ws["C3"].fill=PatternFill("solid",fgColor=GRAY)
    ws.row_dimensions[3].height=15

    # Column headers
    hdr(4, ["SECTION / LINE ITEM","HOURS","RATE × QTY","NOTES"], DARK, WHITE)
    ws.row_dimensions[4].height=18

    r = 5
    def section(title):
        nonlocal r
        ws.merge_cells(f"A{r}:D{r}")
        c=ws.cell(row=r,column=1,value=f"  {title}")
        c.font=Font(name="Arial",bold=True,color=WHITE,size=9)
        c.fill=PatternFill("solid",fgColor=RED)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        ws.row_dimensions[r].height=16
        r+=1

    def data_row(label, hrs, formula="", note=""):
        nonlocal r
        thin=Side(style="thin",color="E0D0D0"); bdr=Border(bottom=thin,left=thin,right=thin,top=thin)
        fill=PatternFill("solid",fgColor=GRAY if r%2==0 else WHITE)
        for ci,v in enumerate([label, f"{hrs:.2f}" if isinstance(hrs,(int,float)) else hrs, formula, note],1):
            c=ws.cell(row=r,column=ci,value=v)
            c.fill=fill; c.border=bdr
            c.font=Font(name="Arial",size=9,color=DARK)
            c.alignment=Alignment(horizontal="right" if ci==2 else "left",vertical="center")
            if ci==2: c.number_format='0.00'
        ws.row_dimensions[r].height=15
        r+=1

    # Use the exact rates that produced this calculation (from _rates_snapshot)
    # so line-item hours match the section totals perfectly
    rates = calc.get("_rates_snapshot") or get_labor_rates()
    bd    = breakdown

    enc_qty = max(1,int(bid.get("enc_qty",1)))
    din_runs= int(bid.get("din_rail_runs",3)); duct_runs=int(bid.get("wire_duct_runs",4))
    wire_cnt= bd.get("wire_count_used",0)
    tb_std=int(bid.get("tb_standard",0)); tb_gnd=int(bid.get("tb_ground",0))
    tb_fsd=int(bid.get("tb_fused",0));    tb_dis=int(bid.get("tb_disconnect",0))
    tb_total=tb_std+tb_gnd+tb_fsd+tb_dis
    di=int(bid.get("plc_di",0)); do_pts=int(bid.get("plc_do",0))
    ai=int(bid.get("plc_ai",0)); ao=int(bid.get("plc_ao",0))
    relay_ic=int(bid.get("relay_icecube",0))
    plc_yn=str(bid.get("plc_present","N")).upper()=="Y"
    hmi_yn=str(bid.get("hmi_present","N")).upper()=="Y"
    eth_sw=1 if str(bid.get("eth_switch","N")).upper()=="Y" else 0
    eth_cab=int(bid.get("eth_cables",0))
    main_amp=int(bid.get("main_amp",100))
    br_1p=int(bid.get("branch_1p",0)); br_2p=int(bid.get("branch_2p",0)); br_3p=int(bid.get("branch_3p",0))
    fused_d=int(bid.get("fused_disconnects",0))
    estops=int(bid.get("estops",0)); pilots=int(bid.get("pilot_lights",0))
    markers=str(bid.get("terminal_markers","Y")).upper()=="Y"
    ferrules=str(bid.get("ferrules","Y")).upper()=="Y"
    hs_yn=str(bid.get("heat_shrink","Y")).upper()=="Y"

    section("ENCLOSURE & MECHANICAL")
    data_row("Enclosure prep + subpanel mount", (rates["enclosure_prep"]+rates["subpanel_mount"])*enc_qty/60,
             f"({rates['enclosure_prep']}+{rates['subpanel_mount']}) min × {enc_qty} enc")
    data_row("Panel layout", rates["panel_layout"]*enc_qty/60, f"{rates['panel_layout']} min × {enc_qty} enc")
    data_row("DIN rail installation", rates["din_rail"]*din_runs/60, f"{rates['din_rail']} min × {din_runs} runs")
    data_row("Wire duct installation", rates["wire_duct"]*duct_runs*4/60, f"{rates['wire_duct']} min × {duct_runs*4} sections")
    data_row("Enclosure accessories", rates["enc_accessory"]*int(bid.get("enc_accessories",0))/60,
             f"{rates['enc_accessory']} min × {bid.get('enc_accessories',0)} items")
    data_row("► SECTION TOTAL", bd.get("enclosure_hrs",0))

    section("POWER DISTRIBUTION")
    mb_rate = rates["main_breaker_small"] if main_amp<=100 else rates["main_breaker_large"]
    data_row("Main breaker", mb_rate*enc_qty/60, f"{mb_rate} min × {enc_qty}")
    data_row("1-pole branch breakers", rates["branch_breaker_1p"]*br_1p/60, f"{rates['branch_breaker_1p']} min × {br_1p}")
    data_row("2/3-pole branch breakers", rates["branch_breaker_23p"]*(br_2p+br_3p)/60, f"{rates['branch_breaker_23p']} min × {br_2p+br_3p}")
    data_row("Fused disconnects", rates["fused_disconnect"]*fused_d/60, f"{rates['fused_disconnect']} min × {fused_d}")
    data_row("► SECTION TOTAL", bd.get("power_hrs",0))

    section("MOTOR CONTROL")
    vfd_sm  = int(bid.get("vfd_small",0))
    vfd_md  = int(bid.get("vfd_med",0))
    vfd_lg  = int(bid.get("vfd_large",0))
    relay_dn= int(bid.get("relay_din",0))
    cont_sm = int(bid.get("contactor_small",0))
    cont_lg = int(bid.get("contactor_large",0))
    overload= int(bid.get("overload",0))
    timer_q = int(bid.get("timers",0))
    ssr_q   = int(bid.get("ssrs",0))
    ss_sm   = int(bid.get("soft_starter_small",0))
    ss_lg   = int(bid.get("soft_starter_large",0))
    if relay_ic:  data_row(f"Ice cube relays ({relay_ic})",   rates["relay_icecube"]*relay_ic/60,   f"{rates['relay_icecube']} min × {relay_ic}")
    if relay_dn:  data_row(f"DIN relays ({relay_dn})",        rates["relay_din"]*relay_dn/60,       f"{rates['relay_din']} min × {relay_dn}")
    if cont_sm:   data_row(f"Contactors ≤40A ({cont_sm})",    rates["contactor_small"]*cont_sm/60,  f"{rates['contactor_small']} min × {cont_sm}")
    if cont_lg:   data_row(f"Contactors >40A ({cont_lg})",    rates["contactor_large"]*cont_lg/60,  f"{rates['contactor_large']} min × {cont_lg}")
    if overload:  data_row(f"Overload relays ({overload})",   rates["overload"]*overload/60,        f"{rates['overload']} min × {overload}")
    if timer_q:   data_row(f"Timer relays ({timer_q})",       rates["timer"]*timer_q/60,            f"{rates['timer']} min × {timer_q}")
    if ssr_q:     data_row(f"Solid state relays ({ssr_q})",   rates["ssr"]*ssr_q/60,               f"{rates['ssr']} min × {ssr_q}")
    if vfd_sm:    data_row(f"VFD ≤5HP ({vfd_sm})",           rates["vfd_small"]*vfd_sm/60,         f"{rates['vfd_small']} min × {vfd_sm}")
    if vfd_md:    data_row(f"VFD 6–25HP ({vfd_md})",         rates["vfd_med"]*vfd_md/60,           f"{rates['vfd_med']} min × {vfd_md}")
    if vfd_lg:    data_row(f"VFD 26–100HP ({vfd_lg})",       rates["vfd_large"]*vfd_lg/60,         f"{rates['vfd_large']} min × {vfd_lg}")
    if ss_sm:     data_row(f"Soft starters ≤50A ({ss_sm})",  rates["soft_starter_small"]*ss_sm/60, f"{rates['soft_starter_small']} min × {ss_sm}")
    if ss_lg:     data_row(f"Soft starters >50A ({ss_lg})",  rates["soft_starter_large"]*ss_lg/60, f"{rates['soft_starter_large']} min × {ss_lg}")
    if not any([relay_ic,relay_dn,cont_sm,cont_lg,overload,timer_q,ssr_q,vfd_sm,vfd_md,vfd_lg,ss_sm,ss_lg]):
        data_row("(no motor control components)", 0)
    data_row("► SECTION TOTAL", bd.get("motor_ctrl_hrs",0))

    section("CONTROL DEVICES")
    data_row("E-stops", rates["estop"]*estops/60, f"{rates['estop']} min × {estops}")
    data_row("Pilot lights", rates["pilot_light"]*pilots/60, f"{rates['pilot_light']} min × {pilots}")
    data_row("► SECTION TOTAL", bd.get("control_dev_hrs",0))

    section("PLC / NETWORKING")
    data_row("PLC rack/controller", (rates["plc_rack"] if plc_yn else 0)/60, f"{rates['plc_rack']} min (if PLC present)")
    data_row("Digital I/O wiring", rates["plc_di_do"]*(di+do_pts)/60, f"{rates['plc_di_do']} min × {di+do_pts} pts (DI:{di} DO:{do_pts})")
    data_row("Analog I/O wiring", rates["plc_ai_ao"]*(ai+ao)/60, f"{rates['plc_ai_ao']} min × {ai+ao} pts (AI:{ai} AO:{ao})")
    data_row("HMI mount & cable", rates["hmi"]*(1 if hmi_yn else 0)/60, f"{rates['hmi']} min (if HMI present)")
    data_row("Ethernet switch", rates["eth_switch"]*eth_sw/60, f"{rates['eth_switch']} min × {eth_sw}")
    data_row("Ethernet cables", rates["eth_cable"]*eth_cab/60, f"{rates['eth_cable']} min × {eth_cab}")
    data_row("► SECTION TOTAL", bd.get("plc_network_hrs",0))

    section("TERMINAL BLOCKS & WIRING")
    data_row("Standard TBs", rates["tb_standard"]*tb_std/60, f"{rates['tb_standard']} min × {tb_std}")
    data_row("Ground TBs", rates["tb_ground"]*tb_gnd/60, f"{rates['tb_ground']} min × {tb_gnd}")
    data_row("Fused TBs", rates["tb_fused"]*tb_fsd/60, f"{rates['tb_fused']} min × {tb_fsd}")
    data_row(f"Wire landing — {wire_cnt} wires × 2 ends", rates["wire_land_control"]*wire_cnt*2/60,
             f"{rates['wire_land_control']} min × {wire_cnt*2} ends", "AUTO-ESTIMATED wire count" if int(bid.get("wire_count",0))==0 else "")
    data_row(f"Ferrule crimping (if enabled)", (rates["ferrule"]*wire_cnt*2/60) if ferrules else 0,
             f"{rates['ferrule']} min × {wire_cnt*2} ends")
    data_row("Wire routing through duct", rates["wire_route"]*wire_cnt/60, f"{rates['wire_route']} min × {wire_cnt}")
    marker_strips=max(1,int(tb_total*1.2/10)) if markers else 0
    data_row("Terminal marker labeling", rates["terminal_markers"]*marker_strips/60, f"{rates['terminal_markers']} min × {marker_strips} strips")
    data_row("► SECTION TOTAL", bd.get("terminals_wire_hrs",0))

    section("HEAT SHRINK LABELS")
    data_row(f"Label application ({wire_cnt*2} labels)", (rates["heat_shrink_label"]*wire_cnt*2/60) if hs_yn else 0,
             f"{rates['heat_shrink_label']} min × {wire_cnt*2}")
    data_row("Heat shrink batch setups", (rates["heat_shrink_batch"]*max(1,int(wire_cnt*2/50))/60) if hs_yn else 0,
             f"{rates['heat_shrink_batch']} min × {max(1,int(wire_cnt*2/50))} batches")
    data_row("► SECTION TOTAL", bd.get("heat_shrink_hrs",0))

    section("UL LABELING & QC")
    data_row("UL component labeling", rates["ul_labels"]*enc_qty/60, f"{rates['ul_labels']} min × {enc_qty} enc")
    data_row("Continuity check", rates["continuity_check"]*wire_cnt/60, f"{rates['continuity_check']} min × {wire_cnt} wires")
    data_row("Hi-pot test", rates["hipot"]*enc_qty/60, f"{rates['hipot']} min × {enc_qty} enc")
    data_row("As-built drawings", rates["as_built"]*enc_qty/60, f"{rates['as_built']} min × {enc_qty} enc")
    data_row("QC sign-off", rates["qc_signoff"]*enc_qty/60, f"{rates['qc_signoff']} min × {enc_qty} enc")
    data_row("► SECTION TOTAL", bd.get("ul_qc_hrs",0))

    # Summary block
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    sc = ws.cell(row=r, column=1, value="  SUMMARY")
    sc.fill = PatternFill("solid", fgColor=DARK_RED)
    sc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 18
    r += 1

    fat_hrs  = float(bid.get("fat_hours", 0))
    eng_hrs  = float(bid.get("eng_hours", 0))
    prog_hrs = float(bid.get("prog_hours", 0))
    data_row("RAW HOURS (sum of all sections)", calc.get("raw_hours", 0))
    data_row(f"Complexity multiplier ({bid.get('complexity','STANDARD')})",
             calc.get("raw_hours",0) * (calc.get("complexity_mult",1.0) - 1.0),
             f"x {calc.get('complexity_mult',1.0):.2f}")
    data_row(f"Tech level multiplier ({bid.get('tech_level','JOURNEYMAN')})", 0,
             f"x {calc.get('tech_mult',1.0):.2f}")
    data_row("Calibration factor", 0, f"x {calc.get('calib_factor',1.0):.3f}")
    data_row("FAT / witness testing hours", fat_hrs)
    data_row("Engineering hours", eng_hrs)
    data_row("Programming hours", prog_hrs)

    # Total row — no merge conflicts: A=label, B=empty, C=value, D=note
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    tc = ws.cell(row=r, column=1, value="  TOTAL BILLABLE HOURS")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=12)
    tc.fill = PatternFill("solid", fgColor=RED)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tv = ws.cell(row=r, column=3, value=calc.get("total_hours", 0))
    tv.font = Font(name="Arial", bold=True, color=WHITE, size=12)
    tv.fill = PatternFill("solid", fgColor=RED)
    tv.alignment = Alignment(horizontal="center", vertical="center")
    tv.number_format = '0.00'
    td = ws.cell(row=r, column=4,
                 value=f"@ ${calc.get('labor_rate_used',95):.0f}/hr = ${calc.get('labor_cost',0):,.2f}")
    td.font = Font(name="Arial", bold=True, color="F4A9A8", size=10)
    td.fill = PatternFill("solid", fgColor=RED)
    ws.row_dimensions[r].height = 24

    ws.freeze_panes="A5"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"AAE_HourBreakdown_{project.replace(' ','_')}.xlsx"
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,download_name=fname)


# ── BOM Converter Routes ─────────────────────────────────────────────────────

@app.route("/api/bom_convert", methods=["POST"])
@require_auth
def bom_convert():
    """Upload one or more PDF files (QuickBooks BOMs or vendor quotes).
    Each PDF is parsed by Claude to extract panels and line items.
    Multiple files are merged into a single response."""
    print("=== /api/bom_convert called ===", flush=True)

    files = request.files.getlist("files")
    if not files or len(files) == 0:
        # Also check single-file upload key
        if "file" in request.files:
            files = [request.files["file"]]
        else:
            return jsonify({"error": "No files uploaded"}), 400

    all_panels = []
    total_tokens = 0
    models_used = set()
    all_flags = []
    file_errors = []

    for f in files:
        pdf_bytes = f.read()
        if len(pdf_bytes) > MAX_SCAN_SIZE:
            file_errors.append(f"{f.filename}: File too large (max 25 MB)")
            continue

        if not pdf_bytes.startswith(PDF_MAGIC):
            file_errors.append(f"{f.filename}: Only PDF files are accepted")
            continue

        # Audit the upload
        file_hash = hashlib.sha256(pdf_bytes).hexdigest()[:16]
        audit_log("bom_convert_upload", "bom_file", file_hash, {
            "filename": f.filename,
            "size_bytes": len(pdf_bytes),
        })

        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
        result = convert_bom_pdf(pdf_b64, f.filename)

        if "error" in result:
            file_errors.append(f"{f.filename}: {result.get('error_message', result['error'])}")
            continue

        panels = result.get("panels", [])
        # Tag each panel with source filename
        for p in panels:
            p["_source_file"] = f.filename
        all_panels.extend(panels)

        if result.get("_model_used"):
            models_used.add(result["_model_used"])
        total_tokens += result.get("_output_tokens", 0)
        all_flags.extend(result.get("extraction_summary", {}).get("review_flags", []))

    # Deduplicate panel names (append suffix if same name appears from different files)
    name_counts = {}
    for p in all_panels:
        name = p["panel_name"]
        if name in name_counts:
            name_counts[name] += 1
            # Only rename duplicates from different source files
            if name_counts[name] > 1:
                p["panel_name"] = f"{name} ({name_counts[name]})"
        else:
            name_counts[name] = 1

    if file_errors:
        all_flags.extend(file_errors)

    # ── Manufacturer Enrichment Pass ─────────────────────────────────────
    # Run all extracted panels through prefix rules + AI to fill in missing manufacturers
    if all_panels:
        try:
            enriched = enrich_manufacturers(all_panels)
            if enriched > 0:
                print(f"BOM_CONVERT: Enriched {enriched} manufacturer fields", flush=True)
                all_flags.append(f"Auto-identified {enriched} manufacturer(s) from part numbers")
        except Exception as e:
            print(f"BOM_CONVERT: Manufacturer enrichment error (non-fatal): {e}", flush=True)

    total_items = sum(len(p.get("line_items", [])) for p in all_panels)

    response_data = {
        "panels": all_panels,
        "extraction_summary": {
            "total_panels_found": len(all_panels),
            "total_line_items": total_items,
            "confidence": 0.9 if all_panels else 0.0,
            "review_flags": all_flags,
            "files_processed": len(files) - len(file_errors),
            "files_errored": len(file_errors),
        },
        "_models_used": list(models_used),
        "_total_output_tokens": total_tokens,
    }

    if not all_panels and file_errors:
        response_data["error"] = "all_files_failed"
        response_data["error_message"] = "; ".join(file_errors)

    return jsonify(response_data)


@app.route("/api/bom_convert_excel", methods=["POST"])
@require_auth
def bom_convert_excel():
    """Generate a ZIP containing Master BOM + per-vendor Excel files for each panel.
    Input: { panels: [...], customer_name, project_name, job_number }
    Output: ZIP file download with folder-per-panel structure."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict
    import zipfile

    data       = request.get_json()
    panels     = data.get("panels", [])
    customer   = data.get("customer_name", "")
    project    = data.get("project_name", "")
    job_num    = data.get("job_number", f"AAE-{datetime.now().strftime('%Y%m%d')}")

    if not panels:
        return jsonify({"error": "No panels to export"}), 400

    routing_rules = load_routing_rules()

    RED      = "9B1B1B"; DARK_RED = "6B0A0A"; WHITE = "FFFFFF"
    LIGHT_RED= "FDECEA"; MID_GRAY = "F5F0F0"; DARK  = "2C2C2C"; TEAL = "00897A"

    def s(cell, bold=False, bg=None, fg=WHITE, sz=10, ha="left", wrap=False):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)

    thin = Side(style="thin", color="E0D0D0")
    bdr  = Border(bottom=thin, left=thin, right=thin, top=thin)

    safe_proj = (project or job_num).replace(" ", "_").replace("/", "-")[:40]

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:

        for panel in panels:
            panel_name = panel.get("panel_name", "Unknown Panel")
            line_items = panel.get("line_items", [])
            if not line_items:
                continue

            safe_panel = panel_name.replace(" ", "_").replace("/", "-").replace("\\", "-")[:40]

            # ── Build Master BOM workbook for this panel ──────────────────────
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Master BOM"

            # 8 columns: Item, Part Number, Description, Qty, U/M, Manufacturer, Vendor, Notes
            for col, w in [("A",8),("B",24),("C",52),("D",8),("E",8),("F",26),("G",18),("H",24)]:
                ws.column_dimensions[col].width = w

            # Row 1: Banner
            ws.merge_cells("A1:H1"); ws.row_dimensions[1].height = 14
            ws["A1"] = "AAE AUTOMATION, INC.  |  UL-NNNY  |  UL-508A Certified Industrial Control Panel Specialists"
            s(ws["A1"], bold=True, bg=RED, sz=11, ha="center")

            # Row 2: Title + Job#
            ws.merge_cells("A2:E2"); ws.row_dimensions[2].height = 34
            ws["A2"] = f"BILL OF MATERIALS — {panel_name}"
            ws["A2"].font      = Font(name="Arial", bold=True, color=WHITE, size=18)
            ws["A2"].fill      = PatternFill("solid", fgColor=DARK_RED)
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells("F2:H2")
            ws["F2"] = job_num
            s(ws["F2"], bold=True, bg=DARK_RED, fg="F4A9A8", sz=12, ha="right")

            # Row 3: Customer + Project
            ws.row_dimensions[3].height = 18
            ws["A3"] = "Customer:"
            s(ws["A3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
            ws.merge_cells("B3:C3"); ws["B3"] = customer
            s(ws["B3"], bg=LIGHT_RED, fg=DARK, sz=10)
            ws["D3"] = "Project:"
            s(ws["D3"], bold=True, bg=LIGHT_RED, fg=DARK_RED, sz=9, ha="right")
            ws.merge_cells("E3:H3"); ws["E3"] = project
            s(ws["E3"], bg=LIGHT_RED, fg=DARK, sz=10)

            # Row 4: Date + Panel Name
            ws.row_dimensions[4].height = 16
            ws["A4"] = "Date:"
            s(ws["A4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
            ws.merge_cells("B4:C4"); ws["B4"] = datetime.now().strftime("%m/%d/%Y")
            s(ws["B4"], bg=MID_GRAY, fg=DARK, sz=9)
            ws["D4"] = "Panel:"
            s(ws["D4"], bold=True, bg=MID_GRAY, fg=DARK_RED, sz=9, ha="right")
            ws.merge_cells("E4:H4"); ws["E4"] = panel_name
            s(ws["E4"], bg=MID_GRAY, fg=DARK, sz=9)

            # Row 5: Column headers
            ws.row_dimensions[5].height = 20
            for ci, h in enumerate(["ITEM","PART NUMBER","DESCRIPTION","QTY","U/M","MANUFACTURER","VENDOR","NOTES"], 1):
                c = ws.cell(row=5, column=ci, value=h)
                s(c, bold=True, bg=DARK, sz=9, ha="center")
                c.border = Border(bottom=Side(style="medium", color=RED))

            # Data rows — chronological order (same order as BOM), no category sections
            row = 6
            even_fill = PatternFill("solid", fgColor="FDF8F8")
            odd_fill  = PatternFill("solid", fgColor=WHITE)

            vendor_groups = defaultdict(list)

            for item_counter, itm in enumerate(line_items, 1):
                fill = even_fill if item_counter % 2 == 0 else odd_fill
                ws.row_dimensions[row].height = 15
                mfr    = itm.get("manufacturer", "")
                pn     = itm.get("part_number", "")
                result = resolve_vendor(pn, mfr, routing_rules)
                vendor = result["vendor"]
                itm["_resolved_vendor"] = vendor

                # Collect for vendor grouping
                norm = dict(itm); norm["vendor"] = vendor
                vendor_groups[vendor].append(norm)

                vals = [
                    item_counter,
                    pn,
                    itm.get("description", ""),
                    itm.get("qty", 1) or 1,
                    itm.get("unit", "ea"),
                    mfr,
                    vendor,
                    itm.get("notes", "")
                ]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.fill = fill; c.border = bdr
                    c.font = Font(name="Arial", size=9, color=DARK)
                    if ci in (1, 4):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    elif ci == 5:  # U/M
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    elif ci == 7:  # Vendor — teal accent
                        c.font = Font(name="Arial", size=9, color=TEAL, bold=bool(val))
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci==3))
                row += 1

            # Total line items row
            ws.merge_cells(f"A{row}:C{row}")
            tl = ws.cell(row=row, column=1, value=f"TOTAL LINE ITEMS: {len(line_items)}")
            tl.font = Font(name="Arial", bold=True, color=DARK_RED, size=10)
            tl.fill = PatternFill("solid", fgColor=LIGHT_RED)
            tl.alignment = Alignment(horizontal="right", vertical="center")
            tv = ws.cell(row=row, column=4, value=sum(int(itm.get("qty") or 1) for itm in line_items))
            tv.font = Font(name="Arial", bold=True, color="008800", size=11)
            tv.fill = PatternFill("solid", fgColor=LIGHT_RED)
            tv.alignment = Alignment(horizontal="center", vertical="center")
            for ci in range(5, 9):
                ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=LIGHT_RED)
            ws.row_dimensions[row].height = 20
            row += 2

            # Footer
            ws.merge_cells(f"A{row}:H{row}")
            ft = ws.cell(row=row, column=1,
                value="AAE Automation, Inc.  |  8528 SW 2nd St, Oklahoma City, OK 73128  |  405-210-1567  |  mfellers@aaeok.com")
            ft.font = Font(name="Arial", size=8, color="888080", italic=True)
            ft.alignment = Alignment(horizontal="center")

            ws.freeze_panes = "A6"
            ws.auto_filter.ref = f"A5:H{row-3}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
            ws.print_title_rows = "1:5"

            # Save master BOM into ZIP
            master_buf = io.BytesIO()
            wb.save(master_buf)
            zf.writestr(f"{safe_panel}/AAE_Master_BOM_{safe_panel}.xlsx", master_buf.getvalue())

            # ── Build per-vendor Excel files for this panel ───────────────────
            for vn in sorted(vendor_groups.keys()):
                if vendor_groups[vn]:
                    vbuf = _build_vendor_excel(vn, vendor_groups[vn],
                                               customer, project or panel_name, job_num, "")
                    safe_vn = vn.replace(" ", "_").replace("/", "-")[:20]
                    zf.writestr(f"{safe_panel}/AAE_BOM_{safe_vn}_{safe_panel}.xlsx", vbuf.getvalue())

    zip_buf.seek(0)
    return send_file(zip_buf,
                     mimetype="application/zip",
                     as_attachment=True,
                     download_name=f"AAE_BOM_Converter_{safe_proj}.zip")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
